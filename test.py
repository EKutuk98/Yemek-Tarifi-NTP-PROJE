import sys
import os
import csv
import json
import shutil
import sqlite3
from datetime import datetime
from contextlib import contextmanager

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_VAR = True
except ImportError:
    OPENPYXL_VAR = False
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem, QDialog, QLabel,
    QLineEdit, QComboBox, QMessageBox, QTabWidget, QFrame, QSpinBox,
    QHeaderView, QGridLayout, QTextEdit, QStackedWidget, QToolBar,
    QCheckBox, QFileDialog, QMenu, QAction, QShortcut, QSplitter,
    QScrollArea, QRadioButton, QButtonGroup, QWizard, QWizardPage,
    QListWidget, QListWidgetItem
)
from PyQt5.QtCore import Qt, QTimer, QPropertyAnimation, QEasingCurve, QPoint
from PyQt5.QtGui import (
    QFont, QColor, QIcon, QPalette, QLinearGradient, QBrush, QPainter,
    QPixmap, QKeySequence
)
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure


# ===================== TIER 2: SAYISAL SIRALANABİLİR TABLO HÜCRESİ =====================

class NumericTableWidgetItem(QTableWidgetItem):
    def __init__(self, text, sayi):
        super().__init__(text)
        self.sayi = sayi

    def __lt__(self, other):
        if isinstance(other, NumericTableWidgetItem):
            return self.sayi < other.sayi
        return super().__lt__(other)


# ===================== VERİTABANI YÖNETİCİSİ =====================

class DatabaseManager:
    def __init__(self, db_name="yemek_tarif.db"):
        self.db_name = db_name
        self.create_tables()
        self._migrate()
        self.ornek_veri_ekle()

    @contextmanager
    def get_connection(self):
        conn = sqlite3.connect(self.db_name)
        conn.row_factory = sqlite3.Row
        try:
            yield conn
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e
        finally:
            conn.close()

    def create_tables(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS tarifler (
                    tarif_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ad TEXT NOT NULL,
                    kategori TEXT NOT NULL,
                    hazirlama_suresi INTEGER NOT NULL,
                    talimat TEXT,
                    zorluk_seviyesi INTEGER DEFAULT 3,
                    is_favorite INTEGER DEFAULT 0,
                    en_son_yapilma DATETIME,
                    goruntu_yolu TEXT,
                    olusturma_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS malzemeler (
                    malzeme_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tarif_id INTEGER NOT NULL,
                    malzeme_adi TEXT NOT NULL,
                    miktar TEXT NOT NULL,
                    birim TEXT DEFAULT 'gr',
                    fiyat REAL DEFAULT 0.0,
                    kalori INTEGER DEFAULT 0,
                    protein REAL DEFAULT 0.0,
                    karbohidrat REAL DEFAULT 0.0,
                    yag REAL DEFAULT 0.0,
                    FOREIGN KEY (tarif_id) REFERENCES tarifler(tarif_id)
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS kullanicilar (
                    kullanici_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ad TEXT NOT NULL,
                    soyad TEXT NOT NULL,
                    email TEXT UNIQUE,
                    kayit_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS degerlendirmeler (
                    degerlendirme_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tarif_id INTEGER NOT NULL,
                    kullanici_id INTEGER NOT NULL,
                    puan INTEGER NOT NULL,
                    yorum TEXT,
                    tarih TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (tarif_id) REFERENCES tarifler(tarif_id),
                    FOREIGN KEY (kullanici_id) REFERENCES kullanicilar(kullanici_id)
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS kategoriler (
                    kategori_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    kategori_adi TEXT UNIQUE NOT NULL
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS yapilmis_log (
                    log_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tarif_id INTEGER NOT NULL,
                    yapilma_tarihi DATETIME DEFAULT CURRENT_TIMESTAMP,
                    notlar TEXT,
                    FOREIGN KEY (tarif_id) REFERENCES tarifler(tarif_id)
                )
            ''')

            # TIER 4: Kişisel Notlar
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS kisisel_notlar (
                    not_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tarif_id INTEGER,
                    baslik TEXT NOT NULL,
                    icerik TEXT,
                    renk TEXT DEFAULT '#f5a623',
                    olusturma_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    guncelleme_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (tarif_id) REFERENCES tarifler(tarif_id)
                )
            ''')

            # TIER 4: Tarif Versiyonları
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS tarif_versiyonlari (
                    versiyon_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tarif_id INTEGER NOT NULL,
                    versiyon_no INTEGER NOT NULL,
                    ad TEXT NOT NULL,
                    kategori TEXT NOT NULL,
                    hazirlama_suresi INTEGER NOT NULL,
                    talimat TEXT,
                    zorluk_seviyesi INTEGER DEFAULT 3,
                    malzemeler_json TEXT,
                    degisiklik_notu TEXT,
                    kayit_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (tarif_id) REFERENCES tarifler(tarif_id)
                )
            ''')

            kategoriler = ['Türk Mutfağı', 'İtalyan Mutfağı', 'Çin Mutfağı', 'Tatlılar',
                           'Çorbalar', 'Salatalar', 'Ana Yemek', 'Kahvaltılık', 'Vegan', 'Glutensiz']
            for kategori in kategoriler:
                cursor.execute('INSERT OR IGNORE INTO kategoriler (kategori_adi) VALUES (?)', (kategori,))

    def _migrate(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()

            cursor.execute("PRAGMA table_info(tarifler)")
            tarif_kolonlari = [row['name'] for row in cursor.fetchall()]
            if 'zorluk_seviyesi' not in tarif_kolonlari:
                cursor.execute('ALTER TABLE tarifler ADD COLUMN zorluk_seviyesi INTEGER DEFAULT 3')
            if 'is_favorite' not in tarif_kolonlari:
                cursor.execute('ALTER TABLE tarifler ADD COLUMN is_favorite INTEGER DEFAULT 0')
            if 'en_son_yapilma' not in tarif_kolonlari:
                cursor.execute('ALTER TABLE tarifler ADD COLUMN en_son_yapilma DATETIME')
            if 'goruntu_yolu' not in tarif_kolonlari:
                cursor.execute('ALTER TABLE tarifler ADD COLUMN goruntu_yolu TEXT')

            cursor.execute("PRAGMA table_info(malzemeler)")
            malzeme_kolonlari = [row['name'] for row in cursor.fetchall()]
            for kolon, tip in [('birim', "TEXT DEFAULT 'gr'"), ('fiyat', 'REAL DEFAULT 0.0'),
                                ('kalori', 'INTEGER DEFAULT 0'), ('protein', 'REAL DEFAULT 0.0'),
                                ('karbohidrat', 'REAL DEFAULT 0.0'), ('yag', 'REAL DEFAULT 0.0')]:
                if kolon not in malzeme_kolonlari:
                    cursor.execute(f'ALTER TABLE malzemeler ADD COLUMN {kolon} {tip}')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS yapilmis_log (
                    log_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tarif_id INTEGER NOT NULL,
                    yapilma_tarihi DATETIME DEFAULT CURRENT_TIMESTAMP,
                    notlar TEXT,
                    FOREIGN KEY (tarif_id) REFERENCES tarifler(tarif_id)
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS kisisel_notlar (
                    not_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tarif_id INTEGER,
                    baslik TEXT NOT NULL,
                    icerik TEXT,
                    renk TEXT DEFAULT '#f5a623',
                    olusturma_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    guncelleme_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (tarif_id) REFERENCES tarifler(tarif_id)
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS tarif_versiyonlari (
                    versiyon_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tarif_id INTEGER NOT NULL,
                    versiyon_no INTEGER NOT NULL,
                    ad TEXT NOT NULL,
                    kategori TEXT NOT NULL,
                    hazirlama_suresi INTEGER NOT NULL,
                    talimat TEXT,
                    zorluk_seviyesi INTEGER DEFAULT 3,
                    malzemeler_json TEXT,
                    degisiklik_notu TEXT,
                    kayit_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (tarif_id) REFERENCES tarifler(tarif_id)
                )
            ''')

    def ornek_veri_ekle(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM tarifler')
            if cursor.fetchone()[0] == 0:
                ornek_tarifler = [
                    ("Mercimek Çorbası", "Çorbalar", 30, "Mercimek, soğan, havuç, salça, baharatlar ile hazırlanır.", 2),
                    ("Mantı", "Türk Mutfağı", 90, "Hamur yoğrulur, küçük karelere kesilir, içine kıymalı harç konur.", 5),
                    ("İskender Kebap", "Türk Mutfağı", 60, "Pide üzerine döner, domates sosu ve yoğurt ile servis edilir.", 4),
                    ("Tiramisu", "İtalyan Mutfağı", 45, "Kahve, mascarpone peyniri, pandispanya ve kakao ile hazırlanır.", 3),
                    ("Çin Usulü Pilav", "Çin Mutfağı", 35, "Pirinç, yumurta, soya sosu, sebzeler ile yapılır.", 2),
                    ("Menemen", "Kahvaltılık", 15, "Yumurta, domates, biber, soğan ile hazırlanır.", 1),
                    ("Baklava", "Tatlılar", 120, "Yufka, ceviz, şerbet ile yapılır.", 5),
                    ("Tarator", "Salatalar", 10, "Ceviz, yoğurt, sarımsak, dereotu ile hazırlanır.", 1),
                    ("Izgara Somon", "Ana Yemek", 25, "Somon fileto, limon, tuz, karabiber ile ızgarada pişirilir.", 3),
                    ("Vegan Burger", "Vegan", 30, "Nohut köftesi, marul, domates, vegan sos ile servis edilir.", 3),
                ]
                for tarif in ornek_tarifler:
                    cursor.execute('''
                        INSERT INTO tarifler (ad, kategori, hazirlama_suresi, talimat, zorluk_seviyesi)
                        VALUES (?, ?, ?, ?, ?)
                    ''', tarif)

                ornek_malzemeler = [
                    # Tarif 1: Mercimek Çorbası
                    (1, "Mercimek Kırmızı", "2", "su bardağı", 15.00, 250, 25.0),
                    (1, "Soğan", "1", "adet", 3.50, 40, 1.0),
                    (1, "Havuç", "1", "adet", 5.00, 35, 0.9),
                    (1, "Tereyağı", "2", "yemek kaşığı", 12.00, 180, 0.2),
                    (1, "Su", "6", "su bardağı", 0.00, 0, 0.0),
                    # Tarif 2: Mantı
                    (2, "Un", "3", "su bardağı", 18.00, 364, 10.0),
                    (2, "Yumurta", "2", "adet", 8.00, 155, 13.0),
                    (2, "Kıyma", "250", "gr", 45.00, 220, 27.0),
                    (2, "Soğan", "1", "adet", 3.50, 40, 1.0),
                    (2, "Tuz", "1", "tatlı kaşığı", 0.00, 0, 0.0),
                    # Tarif 3: İskender Kebap
                    (3, "Döner Eti", "300", "gr", 120.00, 420, 35.0),
                    (3, "Pide Ekmeği", "1", "adet", 15.00, 280, 10.0),
                    (3, "Yoğurt", "200", "ml", 15.00, 130, 3.6),
                    (3, "Domates Sosu", "150", "ml", 10.00, 80, 1.5),
                    (3, "Tereyağı", "1", "yemek kaşığı", 6.00, 90, 0.1),
                    # Tarif 4: Tiramisu
                    (4, "Mascarpone Peyniri", "250", "gr", 50.00, 450, 5.0),
                    (4, "Yumurta Sarısı", "4", "adet", 12.00, 280, 15.0),
                    (4, "Tatlı Şarabı", "100", "ml", 25.00, 100, 0.0),
                    (4, "Kahve", "150", "ml", 12.00, 10, 0.5),
                    (4, "Pandispanya", "300", "gr", 30.00, 320, 4.0),
                    # Tarif 5: Çin Usulü Pilav
                    (5, "Pirinç (Haşlanmış)", "2", "su bardağı", 12.00, 280, 6.0),
                    (5, "Yumurta", "2", "adet", 8.00, 155, 13.0),
                    (5, "Soya Sosu", "3", "yemek kaşığı", 8.00, 30, 4.0),
                    (5, "Sebze Karışımı", "150", "gr", 10.00, 60, 2.0),
                    (5, "Zeytin Yağı", "2", "yemek kaşığı", 20.00, 180, 0.0),
                    # Tarif 6: Menemen
                    (6, "Yumurta", "3", "adet", 12.00, 235, 19.5),
                    (6, "Domates", "2", "adet", 8.00, 36, 1.0),
                    (6, "Yeşil Biber", "1", "adet", 4.00, 30, 1.0),
                    (6, "Soğan", "1", "adet", 3.50, 40, 1.0),
                    (6, "Zeytin Yağı", "2", "yemek kaşığı", 20.00, 180, 0.0),
                    # Tarif 7: Baklava
                    (7, "Yufka", "500", "gr", 20.00, 310, 10.0),
                    (7, "Ceviz", "300", "gr", 50.00, 690, 16.0),
                    (7, "Şeker", "200", "gr", 10.00, 773, 0.0),
                    (7, "Tereyağı", "250", "gr", 150.00, 1800, 0.2),
                    (7, "Su", "1", "su bardağı", 0.00, 0, 0.0),
                    # Tarif 8: Tarator
                    (8, "Ceviz", "100", "gr", 25.00, 230, 5.3),
                    (8, "Yoğurt", "300", "ml", 20.00, 195, 11.0),
                    (8, "Sarımsak", "3", "diş", 2.00, 14, 0.6),
                    (8, "Dereotu", "50", "gr", 5.00, 35, 2.0),
                    (8, "Su", "200", "ml", 0.00, 0, 0.0),
                    # Tarif 9: Izgara Somon
                    (9, "Somon Filesi", "400", "gr", 80.00, 280, 35.0),
                    (9, "Limon", "1", "adet", 3.00, 29, 1.0),
                    (9, "Zeytin Yağı", "2", "yemek kaşığı", 20.00, 180, 0.0),
                    (9, "Tuz ve Karabiber", "1", "çay kaşığı", 0.50, 0, 0.0),
                    (9, "Taze Nane", "10", "gr", 2.00, 3, 0.2),
                    # Tarif 10: Vegan Burger
                    (10, "Nohut", "200", "gr", 12.00, 200, 12.0),
                    (10, "Soğan", "100", "gr", 3.50, 40, 1.0),
                    (10, "Sarımsak", "2", "diş", 1.00, 9, 0.4),
                    (10, "Hamburger Ekmeği", "1", "adet", 8.00, 240, 8.0),
                    (10, "Marul", "100", "gr", 3.00, 15, 1.2),
                    (10, "Domates", "1", "adet", 4.00, 18, 0.9),
                ]
                for m in ornek_malzemeler:
                    cursor.execute('''
                        INSERT INTO malzemeler (tarif_id, malzeme_adi, miktar, birim, fiyat, kalori, protein)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', m)

                ornek_kullanicilar = [
                    ("Ahmet", "Yılmaz", "ahmet@email.com"),
                    ("Ayşe", "Demir", "ayse@email.com"),
                    ("Mehmet", "Kaya", "mehmet@email.com"),
                    ("Zeynep", "Çelik", "zeynep@email.com"),
                ]
                for k in ornek_kullanicilar:
                    cursor.execute('INSERT INTO kullanicilar (ad, soyad, email) VALUES (?, ?, ?)', k)

                # Örnek değerlendirmeler
                ornek_degerlendirmeler = [
                    (1, 1, 5, "Harika bir çorba, kesinlikle tavsiye ederim!"),
                    (1, 2, 4, "Çok lezzetliydi, biraz daha baharatlı yapabilirim."),
                    (2, 1, 5, "Annemin mantısından bile iyiydi!"),
                    (2, 3, 4, "Emek isteyen ama değen bir tarif."),
                    (3, 2, 5, "İskender kebabın tam olması gereken hali!"),
                    (4, 4, 5, "Tiramisu için mükemmel bir tarif!"),
                    (5, 1, 3, "Güzel ama orijinaline yakın değil."),
                    (6, 2, 5, "Kahvaltı için vazgeçilmez tarif!"),
                    (7, 3, 4, "Çıtır çıtır oldu, harika."),
                    (9, 4, 5, "Somon mükemmel pişti!"),
                ]
                for d in ornek_degerlendirmeler:
                    cursor.execute('''
                        INSERT INTO degerlendirmeler (tarif_id, kullanici_id, puan, yorum)
                        VALUES (?, ?, ?, ?)
                    ''', d)

                # Örnek yapılan log
                ornek_yapilanlar = [
                    (1, "İlk deneme, çok beğendim."),
                    (1, "Aileye yaptım, herkese iyi geldi."),
                    (2, "Uzun sürdü ama çok lezzetli."),
                    (6, "10 dakikada hazır!"),
                    (6, "Bu sefer sarımsak ekledim, daha güzel oldu."),
                    (9, "Akşam yemeği için mükemmel."),
                ]
                for y in ornek_yapilanlar:
                    cursor.execute('''
                        INSERT INTO yapilmis_log (tarif_id, notlar)
                        VALUES (?, ?)
                    ''', y)

                # Örnek kişisel notlar
                ornek_notlar = [
                    (1, "Mercimek Çorbası Notu", "Kırmızı mercimek yerine yeşil de olur ama süre uzar.", "#f5a623"),
                    (2, "Mantı İpucu", "Hamuru çok ince açmak kritik, yırtılmaması lazım.", "#4CAF50"),
                    (6, "Menemen Sırrı", "Yumurtayı en sona ekle, tam pişirme!", "#2196F3"),
                    (9, "Somon Alımı", "Taze somon için Balık Pazarı daha ucuz.", "#9c27b0"),
                ]
                for n in ornek_notlar:
                    cursor.execute('''
                        INSERT INTO kisisel_notlar (tarif_id, baslik, icerik, renk)
                        VALUES (?, ?, ?, ?)
                    ''', n)

    # ---------- TARİF İŞLEMLERİ ----------

    def tarif_ekle(self, ad, kategori, hazirlama_suresi, talimat, zorluk_seviyesi=3, goruntu_yolu=None):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO tarifler (ad, kategori, hazirlama_suresi, talimat, zorluk_seviyesi, goruntu_yolu)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (ad, kategori, hazirlama_suresi, talimat, zorluk_seviyesi, goruntu_yolu))
            return cursor.lastrowid

    def tarif_sil(self, tarif_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM malzemeler WHERE tarif_id = ?', (tarif_id,))
            cursor.execute('DELETE FROM degerlendirmeler WHERE tarif_id = ?', (tarif_id,))
            cursor.execute('DELETE FROM yapilmis_log WHERE tarif_id = ?', (tarif_id,))
            cursor.execute('DELETE FROM kisisel_notlar WHERE tarif_id = ?', (tarif_id,))
            cursor.execute('DELETE FROM tarif_versiyonlari WHERE tarif_id = ?', (tarif_id,))
            cursor.execute('DELETE FROM tarifler WHERE tarif_id = ?', (tarif_id,))
            return True

    def tarifleri_getir(self, kategori=None):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            if kategori and kategori != "Tümü":
                cursor.execute('SELECT * FROM tarifler WHERE kategori = ? ORDER BY ad', (kategori,))
            else:
                cursor.execute('SELECT * FROM tarifler ORDER BY ad')
            return [dict(row) for row in cursor.fetchall()]

    def tarif_bul(self, tarif_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM tarifler WHERE tarif_id = ?', (tarif_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    # TIER 4: Tarif güncelleme (snapshot alır)
    def tarif_guncelle(self, tarif_id, ad, kategori, hazirlama_suresi, talimat, zorluk_seviyesi, degisiklik_notu=""):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            # Önce snapshot al
            cursor.execute('SELECT * FROM tarifler WHERE tarif_id = ?', (tarif_id,))
            mevcut = cursor.fetchone()
            if mevcut:
                cursor.execute('SELECT MAX(versiyon_no) as max_v FROM tarif_versiyonlari WHERE tarif_id = ?', (tarif_id,))
                row = cursor.fetchone()
                yeni_v = (row['max_v'] or 0) + 1
                malzemeler = self.malzemeleri_getir(tarif_id)
                cursor.execute('''
                    INSERT INTO tarif_versiyonlari
                    (tarif_id, versiyon_no, ad, kategori, hazirlama_suresi, talimat, zorluk_seviyesi, malzemeler_json, degisiklik_notu)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (tarif_id, yeni_v, mevcut['ad'], mevcut['kategori'], mevcut['hazirlama_suresi'],
                      mevcut['talimat'], mevcut['zorluk_seviyesi'],
                      json.dumps(malzemeler, ensure_ascii=False), degisiklik_notu))
            # Güncelle
            cursor.execute('''
                UPDATE tarifler SET ad=?, kategori=?, hazirlama_suresi=?, talimat=?, zorluk_seviyesi=?
                WHERE tarif_id=?
            ''', (ad, kategori, hazirlama_suresi, talimat, zorluk_seviyesi, tarif_id))

    # TIER 4: Versiyon geçmişi
    def versiyonlari_getir(self, tarif_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT * FROM tarif_versiyonlari WHERE tarif_id = ?
                ORDER BY versiyon_no DESC
            ''', (tarif_id,))
            return [dict(row) for row in cursor.fetchall()]

    def versiyonu_geri_yukle(self, tarif_id, versiyon_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM tarif_versiyonlari WHERE versiyon_id = ?', (versiyon_id,))
            v = cursor.fetchone()
            if v:
                cursor.execute('''
                    UPDATE tarifler SET ad=?, kategori=?, hazirlama_suresi=?, talimat=?, zorluk_seviyesi=?
                    WHERE tarif_id=?
                ''', (v['ad'], v['kategori'], v['hazirlama_suresi'], v['talimat'], v['zorluk_seviyesi'], tarif_id))

    # ---------- KİŞİSEL NOTLAR (TIER 4) ----------

    def not_ekle(self, baslik, icerik, renk='#f5a623', tarif_id=None):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO kisisel_notlar (tarif_id, baslik, icerik, renk)
                VALUES (?, ?, ?, ?)
            ''', (tarif_id, baslik, icerik, renk))
            return cursor.lastrowid

    def not_guncelle(self, not_id, baslik, icerik, renk):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            simdi = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute('''
                UPDATE kisisel_notlar SET baslik=?, icerik=?, renk=?,
                guncelleme_tarihi=? WHERE not_id=?
            ''', (baslik, icerik, renk, simdi, not_id))

    def not_sil(self, not_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM kisisel_notlar WHERE not_id = ?', (not_id,))

    def notlari_getir(self, tarif_id=None):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            if tarif_id:
                cursor.execute('''
                    SELECT n.*, t.ad as tarif_adi FROM kisisel_notlar n
                    LEFT JOIN tarifler t ON n.tarif_id = t.tarif_id
                    WHERE n.tarif_id = ? ORDER BY n.guncelleme_tarihi DESC
                ''', (tarif_id,))
            else:
                cursor.execute('''
                    SELECT n.*, t.ad as tarif_adi FROM kisisel_notlar n
                    LEFT JOIN tarifler t ON n.tarif_id = t.tarif_id
                    ORDER BY n.guncelleme_tarihi DESC
                ''')
            return [dict(row) for row in cursor.fetchall()]

    # ---------- TIER 1: ARAMA ----------

    def tarif_ara(self, arama_terimi):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            arama = f"%{arama_terimi}%"
            cursor.execute('''
                SELECT DISTINCT t.* FROM tarifler t
                LEFT JOIN malzemeler m ON t.tarif_id = m.tarif_id
                WHERE LOWER(t.ad) LIKE LOWER(?) OR LOWER(t.kategori) LIKE LOWER(?)
                   OR LOWER(t.talimat) LIKE LOWER(?) OR LOWER(m.malzeme_adi) LIKE LOWER(?)
                ORDER BY t.ad
            ''', (arama, arama, arama, arama))
            return [dict(row) for row in cursor.fetchall()]

    def tarifleri_sure_araligi(self, min_sure, max_sure):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM tarifler WHERE hazirlama_suresi BETWEEN ? AND ? ORDER BY hazirlama_suresi',
                           (min_sure, max_sure))
            return [dict(row) for row in cursor.fetchall()]

    def favorileri_getir(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM tarifler WHERE is_favorite = 1 ORDER BY ad')
            return [dict(row) for row in cursor.fetchall()]

    def tarif_favoriye_ekle(self, tarif_id, is_favorite=True):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('UPDATE tarifler SET is_favorite = ? WHERE tarif_id = ?',
                           (1 if is_favorite else 0, tarif_id))

    def tarif_zorluk_guncelle(self, tarif_id, zorluk_seviyesi):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('UPDATE tarifler SET zorluk_seviyesi = ? WHERE tarif_id = ?',
                           (zorluk_seviyesi, tarif_id))

    def tarif_resim_guncelle(self, tarif_id, goruntu_yolu):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('UPDATE tarifler SET goruntu_yolu = ? WHERE tarif_id = ?',
                           (goruntu_yolu, tarif_id))

    # ---------- MALİYET & BESİN ----------

    def tahmini_tarif_maliyeti(self, tarif_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT SUM(fiyat) as toplam FROM malzemeler WHERE tarif_id = ?', (tarif_id,))
            row = cursor.fetchone()
            return round(row['toplam'] or 0, 2)

    def tahmini_tarif_kalorileri(self, tarif_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT SUM(kalori) as toplam FROM malzemeler WHERE tarif_id = ?', (tarif_id,))
            row = cursor.fetchone()
            return int(row['toplam'] or 0)

    def tahmini_tarif_proteini(self, tarif_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT SUM(protein) as toplam FROM malzemeler WHERE tarif_id = ?', (tarif_id,))
            row = cursor.fetchone()
            return round(row['toplam'] or 0, 1)

    # ---------- MALZEME ----------

    def malzeme_ekle(self, tarif_id, malzeme_adi, miktar, birim="gr", fiyat=0, kalori=0, protein=0):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO malzemeler (tarif_id, malzeme_adi, miktar, birim, fiyat, kalori, protein)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (tarif_id, malzeme_adi, miktar, birim, fiyat, kalori, protein))
            return cursor.lastrowid

    def malzemeleri_getir(self, tarif_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM malzemeler WHERE tarif_id = ? ORDER BY malzeme_id', (tarif_id,))
            return [dict(row) for row in cursor.fetchall()]

    def malzeme_guncelle(self, malzeme_id, malzeme_adi, miktar, birim, fiyat, kalori, protein):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE malzemeler SET malzeme_adi=?, miktar=?, birim=?, fiyat=?, kalori=?, protein=?
                WHERE malzeme_id=?
            ''', (malzeme_adi, miktar, birim, fiyat, kalori, protein, malzeme_id))

    def malzeme_sil(self, malzeme_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM malzemeler WHERE malzeme_id = ?', (malzeme_id,))

    def tarif_kopyala(self, tarif_id):
        """Tarifi tüm malzemeleriyle birlikte kopyalar."""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM tarifler WHERE tarif_id = ?', (tarif_id,))
            t = cursor.fetchone()
            if not t:
                return None
            cursor.execute('''
                INSERT INTO tarifler (ad, kategori, hazirlama_suresi, talimat, zorluk_seviyesi)
                VALUES (?, ?, ?, ?, ?)
            ''', (f"{t['ad']} (Kopya)", t['kategori'], t['hazirlama_suresi'],
                  t['talimat'], t['zorluk_seviyesi']))
            yeni_id = cursor.lastrowid
            cursor.execute('SELECT * FROM malzemeler WHERE tarif_id = ?', (tarif_id,))
            for m in cursor.fetchall():
                cursor.execute('''
                    INSERT INTO malzemeler (tarif_id, malzeme_adi, miktar, birim, fiyat, kalori, protein)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (yeni_id, m['malzeme_adi'], m['miktar'], m['birim'],
                      m['fiyat'], m['kalori'], m['protein']))
            return yeni_id

    # ---------- KULLANICI ----------

    def kullanici_ekle(self, ad, soyad, email):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('INSERT INTO kullanicilar (ad, soyad, email) VALUES (?, ?, ?)', (ad, soyad, email))
            return cursor.lastrowid

    def kullanici_sil(self, kullanici_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM degerlendirmeler WHERE kullanici_id = ?', (kullanici_id,))
            cursor.execute('DELETE FROM kullanicilar WHERE kullanici_id = ?', (kullanici_id,))
            return True

    def kullanicilari_getir(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM kullanicilar ORDER BY kayit_tarihi DESC')
            return [dict(row) for row in cursor.fetchall()]

    # ---------- DEĞERLENDİRME ----------

    def degerlendirme_ekle(self, tarif_id, kullanici_id, puan, yorum):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            simdi = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute('''
                INSERT INTO degerlendirmeler (tarif_id, kullanici_id, puan, yorum, tarih)
                VALUES (?, ?, ?, ?, ?)
            ''', (tarif_id, kullanici_id, puan, yorum, simdi))
            return cursor.lastrowid

    def degerlendirmeleri_getir(self, tarif_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT d.*, k.ad, k.soyad
                FROM degerlendirmeler d
                JOIN kullanicilar k ON d.kullanici_id = k.kullanici_id
                WHERE d.tarif_id = ? ORDER BY d.tarih DESC
            ''', (tarif_id,))
            return [dict(row) for row in cursor.fetchall()]

    def ortalama_puan(self, tarif_id):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT AVG(puan) as ortalama FROM degerlendirmeler WHERE tarif_id = ?', (tarif_id,))
            row = cursor.fetchone()
            return round(row['ortalama'], 1) if row['ortalama'] else 0

    def kategorileri_getir(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT kategori_adi FROM kategoriler ORDER BY kategori_adi')
            return [row['kategori_adi'] for row in cursor.fetchall()]

    # ---------- TIER 3: YAPILAN LOG ----------

    def yapilmis_log_ekle(self, tarif_id, notlar=""):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            simdi = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute('INSERT INTO yapilmis_log (tarif_id, notlar, yapilma_tarihi) VALUES (?, ?, ?)',
                           (tarif_id, notlar, simdi))
            cursor.execute('UPDATE tarifler SET en_son_yapilma = ? WHERE tarif_id = ?',
                           (simdi, tarif_id))
            return cursor.lastrowid

    def yapilmis_loglari_getir(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT l.*, t.ad FROM yapilmis_log l
                JOIN tarifler t ON l.tarif_id = t.tarif_id
                ORDER BY l.yapilma_tarihi DESC
            ''')
            return [dict(row) for row in cursor.fetchall()]

    def en_cok_yapilan_tarifler(self, limit=5):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT t.ad, COUNT(l.log_id) as sayi FROM tarifler t
                JOIN yapilmis_log l ON t.tarif_id = l.tarif_id
                GROUP BY t.tarif_id ORDER BY sayi DESC LIMIT ?
            ''', (limit,))
            return [dict(row) for row in cursor.fetchall()]

    def zorluk_dagilimi(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT zorluk_seviyesi, COUNT(*) as sayi FROM tarifler GROUP BY zorluk_seviyesi ORDER BY zorluk_seviyesi')
            return [dict(row) for row in cursor.fetchall()]

    def puan_dagilimi(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT puan, COUNT(*) as sayi FROM degerlendirmeler GROUP BY puan ORDER BY puan')
            return [dict(row) for row in cursor.fetchall()]

    def aylik_yapilmis_sayisi(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT strftime('%Y-%m', yapilma_tarihi) as ay, COUNT(*) as sayi
                FROM yapilmis_log GROUP BY ay ORDER BY ay
            ''')
            return [dict(row) for row in cursor.fetchall()]

    def istatistikler(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) as c FROM tarifler')
            toplam_tarif = cursor.fetchone()['c']
            cursor.execute('SELECT COUNT(*) as c FROM kategoriler')
            toplam_kategori = cursor.fetchone()['c']
            cursor.execute('SELECT COUNT(*) as c FROM kullanicilar')
            toplam_kullanici = cursor.fetchone()['c']
            cursor.execute('SELECT COUNT(*) as c FROM degerlendirmeler')
            toplam_degerlendirme = cursor.fetchone()['c']
            cursor.execute('SELECT kategori, COUNT(*) as sayi FROM tarifler GROUP BY kategori ORDER BY sayi DESC')
            kategori_dagilimi = [dict(row) for row in cursor.fetchall()]
            cursor.execute('SELECT COUNT(*) as c FROM tarifler WHERE is_favorite = 1')
            favoriler_sayi = cursor.fetchone()['c']
            cursor.execute('SELECT COUNT(*) as c FROM yapilmis_log')
            toplam_yapilmis = cursor.fetchone()['c']
            cursor.execute('SELECT COUNT(*) as c FROM kisisel_notlar')
            toplam_not = cursor.fetchone()['c']
            cursor.execute('SELECT COUNT(*) as c FROM tarif_versiyonlari')
            toplam_versiyon = cursor.fetchone()['c']
            return {
                "toplam_tarif": toplam_tarif, "toplam_kategori": toplam_kategori,
                "toplam_kullanici": toplam_kullanici, "toplam_degerlendirme": toplam_degerlendirme,
                "kategori_dagilimi": kategori_dagilimi, "favoriler_sayi": favoriler_sayi,
                "toplam_yapilmis": toplam_yapilmis, "toplam_not": toplam_not,
                "toplam_versiyon": toplam_versiyon
            }

    def yedekle(self, hedef_yol):
        shutil.copy2(self.db_name, hedef_yol)
        return True

    def geri_yukle(self, kaynak_yol):
        shutil.copy2(kaynak_yol, self.db_name)
        return True


def dialog_maximize_flag(dialog):
    """QDialog'a minimize/maximize butonları ekler (soru işareti yerine)."""
    dialog.setWindowFlags(
        Qt.Window |
        Qt.WindowTitleHint |
        Qt.WindowSystemMenuHint |
        Qt.WindowMinimizeButtonHint |
        Qt.WindowMaximizeButtonHint |
        Qt.WindowCloseButtonHint
    )


_MSG_STYLE = """
    QMessageBox {
        background-color: #1a1a2e;
        color: #ffffff;
    }
    QMessageBox QLabel {
        color: #ffffff;
        font-size: 13px;
        padding: 6px;
        min-width: 260px;
    }
    QMessageBox QPushButton {
        background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
            stop:0 #f5a623, stop:1 #1a1a2e);
        color: #ffffff;
        border: none;
        border-radius: 8px;
        font-weight: bold;
        font-size: 12px;
        padding: 8px 22px;
        min-width: 80px;
    }
    QPushButton:hover {
        background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
            stop:0 #f5a623, stop:1 #f5a623);
    }
    QPushButton:focus { outline: none; }
"""


def msg_info(parent, baslik, metin):
    mb = QMessageBox(parent)
    mb.setWindowTitle(baslik)
    mb.setText(metin)
    mb.setIcon(QMessageBox.NoIcon)
    mb.setStyleSheet(_MSG_STYLE)
    mb.exec_()


def msg_warn(parent, baslik, metin):
    mb = QMessageBox(parent)
    mb.setWindowTitle(baslik)
    mb.setText(metin)
    mb.setIcon(QMessageBox.NoIcon)
    mb.setStyleSheet(_MSG_STYLE)
    mb.exec_()


def msg_question(parent, baslik, metin):
    mb = QMessageBox(parent)
    mb.setWindowTitle(baslik)
    mb.setText(metin)
    mb.setIcon(QMessageBox.NoIcon)
    mb.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
    mb.setDefaultButton(QMessageBox.No)
    mb.setStyleSheet(_MSG_STYLE)
    # Türkçe buton yazıları
    mb.button(QMessageBox.Yes).setText("✅ Evet")
    mb.button(QMessageBox.No).setText("❌ Hayır")
    return mb.exec_()


# ===================== ÖZEL BUTON SINIFI =====================

class LuxuryButton(QPushButton):
    def __init__(self, text, color="#f5a623", parent=None):
        super().__init__(text, parent)
        self.color = color
        self.setMinimumHeight(36)
        self.setSizePolicy(
            self.sizePolicy().Preferred,
            self.sizePolicy().Fixed
        )
        self.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {color}, stop:1 #1a1a2e);
                color: white; border: none; border-radius: 10px;
                font-weight: bold; font-size: 11px; padding: 7px 12px;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {color}, stop:1 {color});
            }}
        """)

    def enterEvent(self, event):
        self.animation = QPropertyAnimation(self, b"geometry")
        self.animation.setDuration(200)
        self.animation.setEasingCurve(QEasingCurve.OutBounce)
        rect = self.geometry()
        self.animation.setStartValue(rect)
        rect.setWidth(rect.width() + 10)
        self.animation.setEndValue(rect)
        self.animation.start()
        super().enterEvent(event)

    def leaveEvent(self, event):
        self.animation = QPropertyAnimation(self, b"geometry")
        self.animation.setDuration(200)
        rect = self.geometry()
        self.animation.setStartValue(rect)
        rect.setWidth(rect.width() - 10)
        self.animation.setEndValue(rect)
        self.animation.start()
        super().leaveEvent(event)


# ===================== TIER 2: TEMA YÖNETİCİSİ =====================

class ThemeManager:
    DARK = """
        QMainWindow { background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #0f0f1a, stop:1 #1a1a2e); }
        QWidget { background-color: transparent; }
        QLabel { color: #ffffff; }
        QTableWidget { background-color: #2d2d3a; alternate-background-color: #3d3d4a;
            color: #ffffff; gridline-color: #4CAF50; border: none; border-radius: 15px; }
        QTableWidget::item { padding: 12px; color: #ffffff; }
        QTableWidget::item:selected { background-color: #f5a623; color: #1a1a2e; }
        QHeaderView::section { background-color: #1a1a2e; color: #f5a623; font-weight: bold; padding: 12px; border: none; }
        QComboBox, QLineEdit, QSpinBox, QTextEdit { background-color: #1a1a2e; border: 2px solid #f5a623;
            border-radius: 12px; padding: 10px; color: #ffffff; font-size: 12px; }
        QComboBox:focus, QLineEdit:focus { border: 2px solid #4CAF50; }
        QComboBox QAbstractItemView { background-color: #1a1a2e; color: #ffffff; selection-background-color: #f5a623; }
        QCheckBox { color: #f5a623; font-weight: bold; }
        QScrollBar:vertical { background-color: #2d2d3a; border-radius: 10px; width: 10px; }
        QScrollBar::handle:vertical { background-color: #f5a623; border-radius: 10px; }
        QScrollBar::handle:vertical:hover { background-color: #4CAF50; }
        QMenu { background-color: #2d2d3a; color: #ffffff; border: 1px solid #f5a623; border-radius: 8px; }
        QMenu::item:selected { background-color: #f5a623; color: #1a1a2e; }
        QListWidget { background-color: #2d2d3a; color: #ffffff; border: 1px solid #f5a623; border-radius: 10px; }
        QListWidget::item:selected { background-color: #f5a623; color: #1a1a2e; }
        QSplitter::handle { background-color: #f5a623; }
    """
    LIGHT = """
        QMainWindow { background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #f0f2f5, stop:1 #e2e8f0); }
        QWidget { background-color: transparent; }
        QLabel { color: #1a1a2e; }
        QTableWidget { background-color: #ffffff; alternate-background-color: #f1f5f9;
            color: #1a1a2e; gridline-color: #cbd5e1; border: none; border-radius: 15px; }
        QTableWidget::item { padding: 12px; color: #1a1a2e; }
        QTableWidget::item:selected { background-color: #f5a623; color: #ffffff; }
        QHeaderView::section { background-color: #e2e8f0; color: #b8860b; font-weight: bold; padding: 12px; border: none; }
        QComboBox, QLineEdit, QSpinBox, QTextEdit { background-color: #ffffff; border: 2px solid #f5a623;
            border-radius: 12px; padding: 10px; color: #1a1a2e; font-size: 12px; }
        QComboBox:focus, QLineEdit:focus { border: 2px solid #4CAF50; }
        QComboBox QAbstractItemView { background-color: #ffffff; color: #1a1a2e; selection-background-color: #f5a623; }
        QCheckBox { color: #b8860b; font-weight: bold; }
        QScrollBar:vertical { background-color: #e2e8f0; border-radius: 10px; width: 10px; }
        QScrollBar::handle:vertical { background-color: #f5a623; border-radius: 10px; }
        QScrollBar::handle:vertical:hover { background-color: #4CAF50; }
        QMenu { background-color: #ffffff; color: #1a1a2e; border: 1px solid #f5a623; border-radius: 8px; }
        QMenu::item:selected { background-color: #f5a623; color: #ffffff; }
        QListWidget { background-color: #ffffff; color: #1a1a2e; border: 1px solid #f5a623; border-radius: 10px; }
        QListWidget::item:selected { background-color: #f5a623; color: #ffffff; }
        QSplitter::handle { background-color: #f5a623; }
    """

    @staticmethod
    def get(theme):
        return ThemeManager.LIGHT if theme == "light" else ThemeManager.DARK


# ===================== TIER 4: VERSİYON GEÇMİŞİ DİALOGU =====================

class VersiyonGecmisDialog(QDialog):
    def __init__(self, db, tarif_id, parent=None):
        super().__init__(parent)
        self.db = db
        self.tarif_id = tarif_id
        self.tarif = db.tarif_bul(tarif_id)
        self.setWindowTitle(f"📜 Versiyon Geçmişi - {self.tarif['ad']}")
        self.setGeometry(100, 100, 900, 600)
        dialog_maximize_flag(self)
        self.setStyleSheet("""
            QDialog { background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #1a1a2e, stop:1 #2d2d3a); }
            QLabel { color: #ffffff; }
            QListWidget { background-color: #0f0f1a; color: #ffffff; border: 1px solid #f5a623; border-radius: 10px; font-size: 12px; }
            QListWidget::item { padding: 10px; border-bottom: 1px solid #2d2d3a; }
            QListWidget::item:selected { background-color: #f5a623; color: #1a1a2e; }
            QTextEdit { background-color: #0f0f1a; color: #ffffff; border: 1px solid #4CAF50; border-radius: 10px; padding: 15px; font-size: 12px; }
        """)
        self.versiyonlar = db.versiyonlari_getir(tarif_id)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        baslik = QLabel(f"📜 VERSİYON GEÇMİŞİ — {self.tarif['ad']}")
        baslik.setFont(QFont("Arial", 16, QFont.Bold))
        baslik.setStyleSheet("color: #f5a623;")
        baslik.setAlignment(Qt.AlignCenter)
        layout.addWidget(baslik)

        # QSplitter: sol liste, sağ detay
        splitter = QSplitter(Qt.Horizontal)

        # Sol: Versiyon listesi
        sol_widget = QWidget()
        sol_widget.setStyleSheet("background-color: transparent;")
        sol_layout = QVBoxLayout(sol_widget)
        sol_layout.setContentsMargins(0, 0, 0, 0)
        sol_label = QLabel("Versiyonlar:")
        sol_label.setStyleSheet("color: #f5a623; font-weight: bold;")
        sol_layout.addWidget(sol_label)
        self.versiyon_list = QListWidget()
        for v in self.versiyonlar:
            tarih = str(v['kayit_tarihi'])[:16]
            item_text = f"v{v['versiyon_no']}  —  {tarih}\n  {v['degisiklik_notu'][:40] if v['degisiklik_notu'] else 'Not yok'}"
            self.versiyon_list.addItem(item_text)
        self.versiyon_list.currentRowChanged.connect(self.versiyon_sec)
        sol_layout.addWidget(self.versiyon_list)
        splitter.addWidget(sol_widget)

        # Sağ: Versiyon detay
        sag_widget = QWidget()
        sag_widget.setStyleSheet("background-color: transparent;")
        sag_layout = QVBoxLayout(sag_widget)
        sag_layout.setContentsMargins(0, 0, 0, 0)
        sag_label = QLabel("Versiyon Detayı:")
        sag_label.setStyleSheet("color: #f5a623; font-weight: bold;")
        sag_layout.addWidget(sag_label)
        self.detay_text = QTextEdit()
        self.detay_text.setReadOnly(True)
        if not self.versiyonlar:
            self.detay_text.setText("Henüz versiyon geçmişi yok.\n\nTarifı düzenlediğinizde otomatik snapshot alınır.")
        sag_layout.addWidget(self.detay_text)
        splitter.addWidget(sag_widget)

        splitter.setSizes([300, 600])
        layout.addWidget(splitter)

        # Butonlar
        btn_layout = QHBoxLayout()
        self.geri_yukle_btn = LuxuryButton("♻️ VERSİYONU GERİ YÜKLE", "#9c27b0")
        self.geri_yukle_btn.clicked.connect(self.versiyonu_geri_yukle)
        self.geri_yukle_btn.setEnabled(False)
        kapat_btn = LuxuryButton("❌ KAPAT", "#e63946")
        kapat_btn.clicked.connect(self.close)
        btn_layout.addWidget(self.geri_yukle_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(kapat_btn)
        layout.addLayout(btn_layout)

        self.setLayout(layout)
        if self.versiyonlar:
            self.versiyon_list.setCurrentRow(0)

    def versiyon_sec(self, idx):
        if 0 <= idx < len(self.versiyonlar):
            v = self.versiyonlar[idx]
            self.geri_yukle_btn.setEnabled(True)
            metin = f"📌 Versiyon: v{v['versiyon_no']}\n"
            metin += f"📅 Tarih: {str(v['kayit_tarihi'])[:16]}\n"
            metin += f"📝 Değişiklik Notu: {v['degisiklik_notu'] or 'Belirtilmemiş'}\n\n"
            metin += f"─────────────────────\n"
            metin += f"📖 Tarif Adı: {v['ad']}\n"
            metin += f"📚 Kategori: {v['kategori']}\n"
            metin += f"⏱️ Süre: {v['hazirlama_suresi']} dk\n"
            zorluk_map = {1: "Çok Kolay", 2: "Kolay", 3: "Orta", 4: "Zor", 5: "Çok Zor"}
            metin += f"🎯 Zorluk: {zorluk_map.get(v['zorluk_seviyesi'], '?')}\n\n"
            metin += f"👨‍🍳 Talimat:\n{v['talimat'] or 'Yok'}\n\n"
            if v['malzemeler_json']:
                try:
                    malzemeler = json.loads(v['malzemeler_json'])
                    metin += f"🥕 Malzemeler ({len(malzemeler)} adet):\n"
                    for m in malzemeler[:10]:
                        metin += f"  • {m['malzeme_adi']} — {m['miktar']} {m.get('birim','')}\n"
                    if len(malzemeler) > 10:
                        metin += f"  ... ve {len(malzemeler)-10} malzeme daha\n"
                except:
                    pass
            self.detay_text.setText(metin)

    def versiyonu_geri_yukle(self):
        idx = self.versiyon_list.currentRow()
        if 0 <= idx < len(self.versiyonlar):
            v = self.versiyonlar[idx]
            reply = msg_question(self, "Geri Yükle", f"v{v['versiyon_no']} versiyonuna dönmek istediğinize emin misiniz?\n"
                                         "Mevcut tarif üzerine yazılacak.")
            if reply == QMessageBox.Yes:
                self.db.versiyonu_geri_yukle(self.tarif_id, v['versiyon_id'])
                msg_info(self, "Başarılı", f"✅ v{v['versiyon_no']} geri yüklendi!")
                self.accept()


# ===================== TIER 4: KİŞİSEL NOT DİALOGU =====================

class KisiselNotDialog(QDialog):
    NOT_RENKLERI = [
        ("#f5a623", "🟡 Sarı"),
        ("#4CAF50", "🟢 Yeşil"),
        ("#2196F3", "🔵 Mavi"),
        ("#e63946", "🔴 Kırmızı"),
        ("#9c27b0", "🟣 Mor"),
    ]

    def __init__(self, db, mevcut_not=None, tarif_id=None, parent=None):
        super().__init__(parent)
        self.db = db
        self.mevcut_not = mevcut_not
        self.tarif_id = tarif_id
        self.setWindowTitle("📝 Not Ekle" if not mevcut_not else "📝 Notu Düzenle")
        self.setGeometry(200, 200, 520, 480)
        dialog_maximize_flag(self)
        self.setStyleSheet("""
            QDialog { background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #1a1a2e, stop:1 #2d2d3a); }
            QLabel { color: #f5a623; font-weight: bold; font-size: 12px; }
            QLineEdit, QTextEdit { background-color: #0f0f1a; color: #ffffff;
                border: 2px solid #f5a623; border-radius: 10px; padding: 10px; font-size: 12px; }
            QLineEdit:focus, QTextEdit:focus { border: 2px solid #4CAF50; }
            QRadioButton { color: #ffffff; font-size: 13px; padding: 4px; }
        """)
        self.result = None
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(30, 30, 30, 30)

        baslik_label = QLabel("📝 " + ("NOTU DÜZENLE" if self.mevcut_not else "YENİ NOT EKLE"))
        baslik_label.setFont(QFont("Arial", 16, QFont.Bold))
        baslik_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(baslik_label)

        grid = QGridLayout()
        grid.setSpacing(12)

        grid.addWidget(QLabel("Başlık:"), 0, 0)
        self.baslik_input = QLineEdit()
        self.baslik_input.setPlaceholderText("Not başlığını girin...")
        if self.mevcut_not:
            self.baslik_input.setText(self.mevcut_not['baslik'])
        grid.addWidget(self.baslik_input, 0, 1)

        grid.addWidget(QLabel("İçerik:"), 1, 0)
        self.icerik_input = QTextEdit()
        self.icerik_input.setMaximumHeight(140)
        self.icerik_input.setPlaceholderText("Not içeriğini yazın...")
        if self.mevcut_not:
            self.icerik_input.setPlainText(self.mevcut_not['icerik'] or '')
        grid.addWidget(self.icerik_input, 1, 1)

        layout.addLayout(grid)

        # Renk seçimi
        renk_label = QLabel("Renk Seç:")
        layout.addWidget(renk_label)
        self.renk_group = QButtonGroup()
        renk_layout = QHBoxLayout()
        renk_layout.setSpacing(10)
        self.secili_renk = self.mevcut_not['renk'] if self.mevcut_not else '#f5a623'
        for i, (renk_kodu, renk_adi) in enumerate(self.NOT_RENKLERI):
            rb = QRadioButton(renk_adi)
            rb.setStyleSheet(f"QRadioButton {{ color: {renk_kodu}; font-weight: bold; }}")
            rb.setProperty("renk_kodu", renk_kodu)
            if renk_kodu == self.secili_renk:
                rb.setChecked(True)
            self.renk_group.addButton(rb, i)
            renk_layout.addWidget(rb)
        layout.addLayout(renk_layout)

        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(15)
        kaydet_btn = LuxuryButton("✅ KAYDET", "#4CAF50")
        kaydet_btn.clicked.connect(self.kaydet)
        iptal_btn = LuxuryButton("❌ İPTAL", "#e63946")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(kaydet_btn)
        btn_layout.addWidget(iptal_btn)
        layout.addLayout(btn_layout)

        self.setLayout(layout)

    def kaydet(self):
        baslik = self.baslik_input.text().strip()
        if not baslik:
            msg_warn(self, "Hata", "Başlık giriniz!")
            return
        secili_btn = self.renk_group.checkedButton()
        renk = secili_btn.property("renk_kodu") if secili_btn else '#f5a623'
        self.result = {
            'baslik': baslik,
            'icerik': self.icerik_input.toPlainText().strip(),
            'renk': renk
        }
        self.accept()


# ===================== TIER 4: NOTLAR SAYFASI =====================

class NotlarPage(QWidget):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.parent_window = parent
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        # Üst butonlar
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(12)
        ekle_btn = LuxuryButton("➕ YENİ NOT", "#4CAF50")
        ekle_btn.clicked.connect(self.not_ekle)
        yenile_btn = LuxuryButton("🔄 YENİLE", "#2196F3")
        yenile_btn.clicked.connect(self.notlari_yukle)
        btn_layout.addWidget(ekle_btn)
        btn_layout.addWidget(yenile_btn)
        btn_layout.addStretch()
        layout.addWidget(QLabel("📝 Kişisel notlarınız burada. Tarifle ilişkilendirmek için Tarif Detayı'ndan da not ekleyebilirsiniz."))
        layout.addLayout(btn_layout)

        # Scroll area (kart görünümü)
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setStyleSheet("QScrollArea { border: none; background: transparent; }")

        self.scroll_content = QWidget()
        self.scroll_content.setStyleSheet("background: transparent;")
        self.cards_layout = QVBoxLayout(self.scroll_content)
        self.cards_layout.setSpacing(12)
        self.cards_layout.addStretch()

        self.scroll_area.setWidget(self.scroll_content)
        layout.addWidget(self.scroll_area)
        self.setLayout(layout)
        self.notlari_yukle()

    def notlari_yukle(self):
        # Mevcut kartları temizle
        while self.cards_layout.count() > 1:
            item = self.cards_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        notlar = self.db.notlari_getir()
        if not notlar:
            bos = QLabel("📝 Henüz not eklenmemiş.\n\nYukarıdaki '➕ YENİ NOT' butonuyla ilk notunuzu ekleyin!")
            bos.setAlignment(Qt.AlignCenter)
            bos.setStyleSheet("color: #999999; font-size: 14px; padding: 60px;")
            self.cards_layout.insertWidget(self.cards_layout.count() - 1, bos)
        else:
            for not_item in notlar:
                kart = self.not_karti_olustur(not_item)
                self.cards_layout.insertWidget(self.cards_layout.count() - 1, kart)

    def not_karti_olustur(self, not_item):
        renk = not_item.get('renk', '#f5a623')
        kart = QFrame()
        kart.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 15px; border-left: 6px solid {renk};
                padding: 5px;
            }}
        """)

        kart_layout = QVBoxLayout(kart)
        kart_layout.setSpacing(8)
        kart_layout.setContentsMargins(20, 15, 20, 15)

        # Üst satır: başlık + butonlar
        ust_layout = QHBoxLayout()
        baslik_label = QLabel(f"📌 {not_item['baslik']}")
        baslik_label.setFont(QFont("Arial", 13, QFont.Bold))
        baslik_label.setStyleSheet(f"color: {renk};")
        ust_layout.addWidget(baslik_label)
        ust_layout.addStretch()

        if not_item.get('tarif_adi'):
            tarif_label = QLabel(f"🍽️ {not_item['tarif_adi']}")
            tarif_label.setStyleSheet("color: #4CAF50; font-size: 11px;")
            ust_layout.addWidget(tarif_label)

        duzenle_btn = QPushButton("✏️")
        duzenle_btn.setFixedSize(32, 32)
        duzenle_btn.setStyleSheet("""QPushButton { background-color: #2196F3; color: white; border: none;
            border-radius: 8px; font-size: 14px; }
            QPushButton:hover { background-color: #1976D2; }""")
        duzenle_btn.clicked.connect(lambda _, n=not_item: self.not_duzenle(n))
        ust_layout.addWidget(duzenle_btn)

        sil_btn = QPushButton("🗑️")
        sil_btn.setFixedSize(32, 32)
        sil_btn.setStyleSheet("""QPushButton { background-color: #e63946; color: white; border: none;
            border-radius: 8px; font-size: 14px; }
            QPushButton:hover { background-color: #c62828; }""")
        sil_btn.clicked.connect(lambda _, n=not_item: self.not_sil(n))
        ust_layout.addWidget(sil_btn)

        kart_layout.addLayout(ust_layout)

        # İçerik
        if not_item.get('icerik'):
            icerik_label = QLabel(not_item['icerik'][:200] + ("..." if len(not_item['icerik']) > 200 else ""))
            icerik_label.setStyleSheet("color: #cccccc; font-size: 12px;")
            icerik_label.setWordWrap(True)
            kart_layout.addWidget(icerik_label)

        # Alt: tarih
        tarih_label = QLabel(f"🕐 {str(not_item['guncelleme_tarihi'])[:16]}")
        tarih_label.setStyleSheet("color: #777777; font-size: 10px;")
        kart_layout.addWidget(tarih_label)

        return kart

    def not_ekle(self):
        dialog = KisiselNotDialog(self.db, parent=self)
        if dialog.exec_() == QDialog.Accepted and dialog.result:
            self.db.not_ekle(dialog.result['baslik'], dialog.result['icerik'], dialog.result['renk'])
            self.notlari_yukle()

    def not_duzenle(self, not_item):
        dialog = KisiselNotDialog(self.db, mevcut_not=not_item, parent=self)
        if dialog.exec_() == QDialog.Accepted and dialog.result:
            self.db.not_guncelle(not_item['not_id'], dialog.result['baslik'],
                                 dialog.result['icerik'], dialog.result['renk'])
            self.notlari_yukle()

    def not_sil(self, not_item):
        reply = msg_question(self, "Sil", f"'{not_item['baslik']}' notunu silmek istiyor musunuz?")
        if reply == QMessageBox.Yes:
            self.db.not_sil(not_item['not_id'])
            self.notlari_yukle()


# ===================== TIER 4: TARİF DÜZENLEME DİALOGU =====================

class TarifDuzenleDialog(QDialog):
    def __init__(self, db, tarif_id, parent=None):
        super().__init__(parent)
        self.db = db
        self.tarif = db.tarif_bul(tarif_id)
        self.setWindowTitle(f"✏️ Tarif Düzenle - {self.tarif['ad']}")
        self.setGeometry(100, 100, 620, 680)
        dialog_maximize_flag(self)
        self.setStyleSheet("""
            QDialog { background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #1a1a2e, stop:1 #2d2d3a); }
            QLabel { color: #f5a623; font-weight: bold; font-size: 12px; }
            QLineEdit, QComboBox, QSpinBox, QTextEdit { background-color: #0f0f1a; color: #ffffff;
                border: 2px solid #f5a623; border-radius: 12px; padding: 12px; font-size: 12px; }
            QLineEdit:focus, QComboBox:focus, QTextEdit:focus { border: 2px solid #4CAF50; }
        """)
        self.result = None
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(18)
        layout.setContentsMargins(30, 30, 30, 30)

        baslik = QLabel("✏️ TARİF DÜZENLE")
        baslik.setFont(QFont("Arial", 18, QFont.Bold))
        baslik.setStyleSheet("color: #f5a623;")
        baslik.setAlignment(Qt.AlignCenter)
        layout.addWidget(baslik)

        grid = QGridLayout()
        grid.setSpacing(15)

        grid.addWidget(QLabel("📖 Tarif Adı:"), 0, 0)
        self.ad_input = QLineEdit(self.tarif['ad'])
        grid.addWidget(self.ad_input, 0, 1)

        grid.addWidget(QLabel("📚 Kategori:"), 1, 0)
        self.kategori_combo = QComboBox()
        self.kategori_combo.addItems(self.db.kategorileri_getir())
        idx = self.kategori_combo.findText(self.tarif['kategori'])
        if idx >= 0:
            self.kategori_combo.setCurrentIndex(idx)
        grid.addWidget(self.kategori_combo, 1, 1)

        grid.addWidget(QLabel("⏱️ Hazırlama Süresi (dk):"), 2, 0)
        self.sure_input = QSpinBox()
        self.sure_input.setRange(1, 600)
        self.sure_input.setValue(self.tarif['hazirlama_suresi'])
        grid.addWidget(self.sure_input, 2, 1)

        grid.addWidget(QLabel("🎯 Zorluk Seviyesi:"), 3, 0)
        self.zorluk_combo = QComboBox()
        self.zorluk_combo.addItems(["👶 Çok Kolay (1)", "😊 Kolay (2)", "😐 Orta (3)", "🤓 Zor (4)", "🔥 Çok Zor (5)"])
        self.zorluk_combo.setCurrentIndex((self.tarif['zorluk_seviyesi'] or 3) - 1)
        grid.addWidget(self.zorluk_combo, 3, 1)

        grid.addWidget(QLabel("👨‍🍳 Talimat:"), 4, 0)
        self.talimat_input = QTextEdit()
        self.talimat_input.setMaximumHeight(130)
        self.talimat_input.setPlainText(self.tarif['talimat'] or '')
        grid.addWidget(self.talimat_input, 4, 1)

        grid.addWidget(QLabel("📝 Değişiklik Notu:"), 5, 0)
        self.degisiklik_input = QLineEdit()
        self.degisiklik_input.setPlaceholderText("Ne değiştirdiniz? (opsiyonel, versiyona kaydedilir)")
        grid.addWidget(self.degisiklik_input, 5, 1)

        layout.addLayout(grid)

        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(15)
        kaydet_btn = LuxuryButton("✅ KAYDET", "#4CAF50")
        kaydet_btn.clicked.connect(self.kaydet)
        iptal_btn = LuxuryButton("❌ İPTAL", "#e63946")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(kaydet_btn)
        btn_layout.addWidget(iptal_btn)
        layout.addLayout(btn_layout)

        self.setLayout(layout)

    def kaydet(self):
        ad = self.ad_input.text().strip()
        if not ad:
            msg_warn(self, "Hata", "Tarif adı giriniz!")
            return
        self.result = {
            'ad': ad,
            'kategori': self.kategori_combo.currentText(),
            'hazirlama_suresi': self.sure_input.value(),
            'talimat': self.talimat_input.toPlainText().strip(),
            'zorluk_seviyesi': self.zorluk_combo.currentIndex() + 1,
            'degisiklik_notu': self.degisiklik_input.text().strip()
        }
        self.accept()


# ===================== TIER 4: TARIF WIZARD (QWizard, 3 sayfa) =====================

class TarifWizard(QDialog):
    """
    Tamamen özel 3 adımlı tarif ekleme sihirbazı.
    QWizard kullanılmaz — arka planla tam uyumlu dark dashboard teması.
    """
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.malzeme_listesi = []
        self.setWindowTitle("✨ Tarif Ekleme Sihirbazı")
        self.setGeometry(200, 100, 860, 680)
        self.setMinimumSize(800, 600)
        dialog_maximize_flag(self)
        self.setStyleSheet("""
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #0f0f1a, stop:1 #1a1a2e);
            }
            QWidget { background-color: transparent; }
            QLabel { color: #ffffff; }
            QLineEdit, QComboBox, QSpinBox, QTextEdit {
                background-color: #1a1a2e;
                border: 2px solid #f5a623;
                border-radius: 10px;
                padding: 10px;
                color: #ffffff;
                font-size: 12px;
            }
            QLineEdit:focus, QComboBox:focus, QTextEdit:focus {
                border: 2px solid #4CAF50;
            }
            QComboBox QAbstractItemView {
                background-color: #1a1a2e;
                color: #ffffff;
                selection-background-color: #f5a623;
            }
            QSpinBox { color: #ffffff; }
            QRadioButton { color: #ffffff; font-size: 12px; padding: 4px; }
            QListWidget {
                background-color: #1a1a2e;
                color: #ffffff;
                border: 1px solid #f5a623;
                border-radius: 10px;
                font-size: 12px;
            }
            QListWidget::item { padding: 8px; border-bottom: 1px solid #2d2d3a; }
            QListWidget::item:selected { background-color: #f5a623; color: #1a1a2e; }
            QTableWidget {
                background-color: #1a1a2e;
                color: #ffffff;
                border: 1px solid #f5a623;
                border-radius: 10px;
                gridline-color: #2d2d3a;
            }
            QTableWidget::item { padding: 8px; }
            QHeaderView::section {
                background-color: #0f0f1a;
                color: #f5a623;
                font-weight: bold;
                padding: 8px;
                border: none;
            }
            QScrollBar:vertical { background-color: #2d2d3a; border-radius: 6px; width: 8px; }
            QScrollBar::handle:vertical { background-color: #f5a623; border-radius: 6px; }
        """)
        self.mevcut_adim = 0
        self.init_ui()

    # ── ANA YAPI ──────────────────────────────────────────────

    def init_ui(self):
        ana = QVBoxLayout(self)
        ana.setContentsMargins(0, 0, 0, 0)
        ana.setSpacing(0)

        # ── Başlık barı ──
        baslik_bar = QFrame()
        baslik_bar.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #1a1a2e, stop:0.4 #2d2d3a, stop:0.6 #2d2d3a, stop:1 #1a1a2e);
                border-bottom: 2px solid #f5a623;
            }
        """)
        baslik_bar.setFixedHeight(80)
        bb_layout = QHBoxLayout(baslik_bar)
        bb_layout.setContentsMargins(30, 0, 30, 0)

        ikon = QLabel("✨")
        ikon.setFont(QFont("Arial", 28))
        bb_layout.addWidget(ikon)

        bb_texts = QVBoxLayout()
        bb_texts.setSpacing(2)
        baslik_lbl = QLabel("TARİF EKLEME SİHİRBAZI")
        baslik_lbl.setFont(QFont("Arial", 16, QFont.Bold))
        baslik_lbl.setStyleSheet("color: #f5a623; letter-spacing: 2px;")
        alt_lbl = QLabel("Adım adım yeni tarif ekleyin")
        alt_lbl.setStyleSheet("color: #4CAF50; font-size: 11px;")
        bb_texts.addWidget(baslik_lbl)
        bb_texts.addWidget(alt_lbl)
        bb_layout.addLayout(bb_texts)
        bb_layout.addStretch()

        # Adım indikatörleri
        adim_layout = QHBoxLayout()
        adim_layout.setSpacing(8)
        self.adim_lbls = []
        adimlar = ["① Temel Bilgiler", "② Malzemeler", "③ Özet"]
        for i, ad in enumerate(adimlar):
            lbl = QLabel(ad)
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setFixedHeight(32)
            lbl.setFixedWidth(140)
            lbl.setStyleSheet("""
                QLabel {
                    background-color: #2d2d3a;
                    color: #777777;
                    border-radius: 8px;
                    font-size: 11px;
                    font-weight: bold;
                    padding: 2px 6px;
                }
            """)
            adim_layout.addWidget(lbl)
            self.adim_lbls.append(lbl)
            if i < 2:
                sep = QLabel("›")
                sep.setStyleSheet("color: #444444; font-size: 16px;")
                adim_layout.addWidget(sep)
        bb_layout.addLayout(adim_layout)
        ana.addWidget(baslik_bar)

        # ── İçerik alanı (QStackedWidget) ──
        self.stack = QStackedWidget()
        self.stack.setStyleSheet("QStackedWidget { background: transparent; }")
        self.stack.addWidget(self._sayfa1())
        self.stack.addWidget(self._sayfa2())
        self.stack.addWidget(self._sayfa3())
        ana.addWidget(self.stack)

        # ── Alt navigasyon barı ──
        alt_bar = QFrame()
        alt_bar.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #1a1a2e, stop:0.4 #2d2d3a, stop:0.6 #2d2d3a, stop:1 #1a1a2e);
                border-top: 2px solid #f5a623;
            }
        """)
        alt_bar.setFixedHeight(72)
        alt_layout = QHBoxLayout(alt_bar)
        alt_layout.setContentsMargins(30, 0, 30, 0)
        alt_layout.setSpacing(15)

        self.iptal_btn = LuxuryButton("✕ İPTAL", "#e63946")
        self.iptal_btn.setFixedWidth(110)
        self.iptal_btn.clicked.connect(self.reject)

        alt_layout.addWidget(self.iptal_btn)
        alt_layout.addStretch()

        # İlerleme barı
        self.ilerleme_lbl = QLabel("Adım 1 / 3")
        self.ilerleme_lbl.setStyleSheet("color: #999999; font-size: 11px;")
        alt_layout.addWidget(self.ilerleme_lbl)

        self.geri_btn = LuxuryButton("◀ GERİ", "#795548")
        self.geri_btn.setFixedWidth(110)
        self.geri_btn.setEnabled(False)
        self.geri_btn.clicked.connect(self.geri)

        self.ileri_btn = LuxuryButton("İLERİ ▶", "#2196F3")
        self.ileri_btn.setFixedWidth(130)
        self.ileri_btn.clicked.connect(self.ileri)

        self.kaydet_btn = LuxuryButton("✅ KAYDET", "#4CAF50")
        self.kaydet_btn.setFixedWidth(130)
        self.kaydet_btn.setVisible(False)
        self.kaydet_btn.clicked.connect(self.kaydet)

        alt_layout.addWidget(self.geri_btn)
        alt_layout.addWidget(self.ileri_btn)
        alt_layout.addWidget(self.kaydet_btn)
        ana.addWidget(alt_bar)

        self._adim_guncelle()

    # ── SAYFA 1: TEMEL BİLGİLER ───────────────────────────────

    def _sayfa1(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(40, 30, 40, 20)
        layout.setSpacing(20)

        baslik = QLabel("📖 Temel Tarif Bilgileri")
        baslik.setFont(QFont("Arial", 14, QFont.Bold))
        baslik.setStyleSheet("color: #f5a623;")
        layout.addWidget(baslik)

        # Kart
        kart = QFrame()
        kart.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 15px;
                border: 1px solid #3d3d4a;
            }
        """)
        kl = QGridLayout(kart)
        kl.setContentsMargins(30, 25, 30, 25)
        kl.setSpacing(16)
        kl.setColumnMinimumWidth(0, 160)

        def lbl(txt):
            l = QLabel(txt)
            l.setStyleSheet("color: #f5a623; font-weight: bold; font-size: 12px;")
            return l

        # Tarif adı
        kl.addWidget(lbl("📖 Tarif Adı:"), 0, 0)
        self.ad_input = QLineEdit()
        self.ad_input.setPlaceholderText("Tarifin adını girin...")
        kl.addWidget(self.ad_input, 0, 1)

        # Kategori
        kl.addWidget(lbl("📚 Kategori:"), 1, 0)
        self.kategori_combo = QComboBox()
        self.kategori_combo.addItems(self.db.kategorileri_getir())
        kl.addWidget(self.kategori_combo, 1, 1)

        # Süre
        kl.addWidget(lbl("⏱️ Hazırlama Süresi (dk):"), 2, 0)
        self.sure_input = QSpinBox()
        self.sure_input.setRange(1, 600)
        self.sure_input.setValue(30)
        self.sure_input.setStyleSheet("""
            QSpinBox { background-color: #1a1a2e; color: #ffffff;
                border: 2px solid #f5a623; border-radius: 10px; padding: 10px; font-size: 12px; }
            QSpinBox:focus { border: 2px solid #4CAF50; }
            QSpinBox::up-button, QSpinBox::down-button {
                background-color: #2d2d3a; border: none; width: 20px;
            }
        """)
        kl.addWidget(self.sure_input, 2, 1)

        layout.addWidget(kart)

        # Zorluk — ayrı kart
        zorluk_kart = QFrame()
        zorluk_kart.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 15px;
                border: 1px solid #3d3d4a;
            }
        """)
        zl = QVBoxLayout(zorluk_kart)
        zl.setContentsMargins(30, 20, 30, 20)
        zl.setSpacing(14)

        z_baslik = QLabel("🎯 Zorluk Seviyesi:")
        z_baslik.setStyleSheet("color: #f5a623; font-weight: bold; font-size: 12px;")
        zl.addWidget(z_baslik)

        self.zorluk_group = QButtonGroup()
        zorluk_row = QHBoxLayout()
        zorluk_row.setSpacing(0)
        zorluk_items = [
            ("👶", "Çok Kolay", 1, "#4CAF50"),
            ("😊", "Kolay", 2, "#8BC34A"),
            ("😐", "Orta", 3, "#f5a623"),
            ("🤓", "Zor", 4, "#FF9800"),
            ("🔥", "Çok Zor", 5, "#e63946"),
        ]
        self.zorluk_btns = []
        for emoji, etiket, deger, renk in zorluk_items:
            btn = QPushButton(f"{emoji}\n{etiket}")
            btn.setCheckable(True)
            btn.setFixedHeight(64)
            btn.setProperty("deger", deger)
            btn.setProperty("renk", renk)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: #1a1a2e;
                    color: #777777;
                    border: 2px solid #3d3d4a;
                    border-radius: 0px;
                    font-size: 11px;
                    font-weight: bold;
                    padding: 4px;
                }}
                QPushButton:checked {{
                    background-color: {renk};
                    color: #ffffff;
                    border: 2px solid {renk};
                }}
                QPushButton:hover:!checked {{
                    border-color: {renk};
                    color: #cccccc;
                }}
            """)
            if deger == 3:
                btn.setChecked(True)
            self.zorluk_group.addButton(btn, deger)
            zorluk_row.addWidget(btn)
            self.zorluk_btns.append(btn)

        # Köşe yuvarlatma için ilk ve son
        self.zorluk_btns[0].setStyleSheet(self.zorluk_btns[0].styleSheet().replace(
            "border-radius: 0px;", "border-radius: 10px 0px 0px 10px;"))
        self.zorluk_btns[-1].setStyleSheet(self.zorluk_btns[-1].styleSheet().replace(
            "border-radius: 0px;", "border-radius: 0px 10px 10px 0px;"))

        zl.addLayout(zorluk_row)
        layout.addWidget(zorluk_kart)

        # Talimat
        talimat_kart = QFrame()
        talimat_kart.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 15px;
                border: 1px solid #3d3d4a;
            }
        """)
        tl = QVBoxLayout(talimat_kart)
        tl.setContentsMargins(30, 20, 30, 20)
        tl.setSpacing(10)
        t_baslik = QLabel("👨‍🍳 Talimat (Yapılış Adımları):")
        t_baslik.setStyleSheet("color: #f5a623; font-weight: bold; font-size: 12px;")
        tl.addWidget(t_baslik)
        self.talimat_input = QTextEdit()
        self.talimat_input.setPlaceholderText("Tarifin adım adım yapılışını yazın...")
        self.talimat_input.setMinimumHeight(90)
        self.talimat_input.setMaximumHeight(120)
        tl.addWidget(self.talimat_input)
        layout.addWidget(talimat_kart)
        layout.addStretch()
        return w

    # ── SAYFA 2: MALZEMELER ───────────────────────────────────

    def _sayfa2(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(40, 30, 40, 20)
        layout.setSpacing(16)

        baslik = QLabel("🥕 Malzeme Ekle")
        baslik.setFont(QFont("Arial", 14, QFont.Bold))
        baslik.setStyleSheet("color: #f5a623;")
        layout.addWidget(baslik)

        # Giriş kartı
        giris_kart = QFrame()
        giris_kart.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 15px; border: 1px solid #3d3d4a;
            }
        """)
        gl = QGridLayout(giris_kart)
        gl.setContentsMargins(25, 20, 25, 20)
        gl.setSpacing(12)

        def flbl(txt):
            l = QLabel(txt)
            l.setStyleSheet("color: #4CAF50; font-weight: bold; font-size: 11px;")
            return l

        gl.addWidget(flbl("Malzeme Adı"), 0, 0)
        self.w_ad = QLineEdit()
        self.w_ad.setPlaceholderText("Örn: Domates, Un, Yumurta...")
        gl.addWidget(self.w_ad, 1, 0)

        gl.addWidget(flbl("Miktar"), 0, 1)
        self.w_miktar = QLineEdit()
        self.w_miktar.setPlaceholderText("Örn: 2, 250")
        self.w_miktar.setMaximumWidth(120)
        gl.addWidget(self.w_miktar, 1, 1)

        gl.addWidget(flbl("Birim"), 0, 2)
        self.w_birim = QComboBox()
        self.w_birim.addItems(["gr", "ml", "adet", "su bardağı",
                                "yemek kaşığı", "tatlı kaşığı", "çay kaşığı",
                                "kase", "fincan", "diş", "paket", "demet"])
        self.w_birim.setMaximumWidth(140)
        gl.addWidget(self.w_birim, 1, 2)

        gl.addWidget(flbl("Fiyat (₺)"), 0, 3)
        self.w_fiyat = QSpinBox()
        self.w_fiyat.setRange(0, 10000)
        self.w_fiyat.setStyleSheet("""
            QSpinBox { background-color: #1a1a2e; color: #ffffff;
                border: 2px solid #f5a623; border-radius: 10px; padding: 10px; font-size: 12px; }
            QSpinBox::up-button, QSpinBox::down-button { background-color: #2d2d3a; border: none; width: 18px; }
        """)
        self.w_fiyat.setMaximumWidth(110)
        gl.addWidget(self.w_fiyat, 1, 3)

        gl.addWidget(flbl("Kalori"), 0, 4)
        self.w_kalori = QSpinBox()
        self.w_kalori.setRange(0, 10000)
        self.w_kalori.setStyleSheet(self.w_fiyat.styleSheet())
        self.w_kalori.setMaximumWidth(110)
        gl.addWidget(self.w_kalori, 1, 4)

        ekle_btn = LuxuryButton("➕ EKLE", "#4CAF50")
        ekle_btn.setFixedHeight(42)
        ekle_btn.clicked.connect(self._malzeme_ekle)
        gl.addWidget(ekle_btn, 1, 5)

        layout.addWidget(giris_kart)

        # Malzeme listesi kartı
        liste_kart = QFrame()
        liste_kart.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 15px; border: 1px solid #3d3d4a;
            }
        """)
        ll = QVBoxLayout(liste_kart)
        ll.setContentsMargins(20, 15, 20, 15)
        ll.setSpacing(10)

        l_ust = QHBoxLayout()
        l_baslik = QLabel("🥕 Eklenen Malzemeler")
        l_baslik.setStyleSheet("color: #4CAF50; font-weight: bold; font-size: 12px;")
        l_ust.addWidget(l_baslik)
        l_ust.addStretch()
        self.malzeme_sayisi_lbl = QLabel("0 malzeme")
        self.malzeme_sayisi_lbl.setStyleSheet("color: #777777; font-size: 11px;")
        l_ust.addWidget(self.malzeme_sayisi_lbl)
        sil_btn = LuxuryButton("🗑️ Seçiliyi Sil", "#e63946")
        sil_btn.setMaximumHeight(32)
        sil_btn.clicked.connect(self._malzeme_sil)
        l_ust.addWidget(sil_btn)
        ll.addLayout(l_ust)

        self.w_malzeme_table = QTableWidget()
        self.w_malzeme_table.setColumnCount(5)
        self.w_malzeme_table.setHorizontalHeaderLabels(
            ["Malzeme", "Miktar", "Birim", "Fiyat (₺)", "Kalori"])
        self.w_malzeme_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.w_malzeme_table.setAlternatingRowColors(True)
        self.w_malzeme_table.setStyleSheet("""
            QTableWidget { alternate-background-color: #2d2d3a; background-color: #1a1a2e; }
        """)
        self.w_malzeme_table.setMinimumHeight(200)
        ll.addWidget(self.w_malzeme_table)

        # Toplam özet
        self.malzeme_ozet_lbl = QLabel("💰 Tahmini Maliyet: ₺0.00  |  🔥 Toplam Kalori: 0 cal")
        self.malzeme_ozet_lbl.setStyleSheet(
            "color: #f5a623; font-weight: bold; font-size: 11px; padding: 4px;")
        ll.addWidget(self.malzeme_ozet_lbl)

        layout.addWidget(liste_kart)
        layout.addStretch()
        return w

    # ── SAYFA 3: ÖZET ─────────────────────────────────────────

    def _sayfa3(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(40, 30, 40, 20)
        layout.setSpacing(16)

        baslik = QLabel("✅ Özet & Onay")
        baslik.setFont(QFont("Arial", 14, QFont.Bold))
        baslik.setStyleSheet("color: #f5a623;")
        layout.addWidget(baslik)

        # Özet kart
        ozet_kart = QFrame()
        ozet_kart.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 15px;
                border: 1px solid #f5a623;
            }
        """)
        ol = QVBoxLayout(ozet_kart)
        ol.setContentsMargins(30, 25, 30, 25)
        ol.setSpacing(14)

        # Bilgi satırları
        self.ozet_info_layout = QGridLayout()
        self.ozet_info_layout.setSpacing(12)
        self.ozet_info_layout.setColumnMinimumWidth(0, 160)
        ol.addLayout(self.ozet_info_layout)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("background-color: #3d3d4a; max-height: 1px; margin: 4px 0;")
        ol.addWidget(sep)

        # Malzeme özet tablosu
        malzeme_lbl = QLabel("🥕 Malzemeler:")
        malzeme_lbl.setStyleSheet("color: #4CAF50; font-weight: bold; font-size: 12px;")
        ol.addWidget(malzeme_lbl)

        self.ozet_malzeme_list = QListWidget()
        self.ozet_malzeme_list.setMaximumHeight(160)
        self.ozet_malzeme_list.setStyleSheet("""
            QListWidget { background-color: #0f0f1a; border: 1px solid #2d2d3a; }
        """)
        ol.addWidget(self.ozet_malzeme_list)

        # Besin özeti
        self.ozet_besin_lbl = QLabel("")
        self.ozet_besin_lbl.setStyleSheet(
            "color: #f5a623; font-weight: bold; font-size: 12px; padding: 6px;")
        self.ozet_besin_lbl.setAlignment(Qt.AlignCenter)
        ol.addWidget(self.ozet_besin_lbl)

        layout.addWidget(ozet_kart)
        layout.addStretch()
        return w

    # ── NAVİGASYON ────────────────────────────────────────────

    def _adim_guncelle(self):
        """Adım göstergelerini ve butonları güncelle."""
        for i, lbl in enumerate(self.adim_lbls):
            if i == self.mevcut_adim:
                lbl.setStyleSheet("""
                    QLabel {
                        background-color: #f5a623;
                        color: #1a1a2e;
                        border-radius: 8px;
                        font-size: 11px;
                        font-weight: bold;
                        padding: 2px 6px;
                    }
                """)
            elif i < self.mevcut_adim:
                lbl.setStyleSheet("""
                    QLabel {
                        background-color: #4CAF50;
                        color: #ffffff;
                        border-radius: 8px;
                        font-size: 11px;
                        font-weight: bold;
                        padding: 2px 6px;
                    }
                """)
            else:
                lbl.setStyleSheet("""
                    QLabel {
                        background-color: #2d2d3a;
                        color: #777777;
                        border-radius: 8px;
                        font-size: 11px;
                        font-weight: bold;
                        padding: 2px 6px;
                    }
                """)

        self.stack.setCurrentIndex(self.mevcut_adim)
        self.geri_btn.setEnabled(self.mevcut_adim > 0)
        self.ilerleme_lbl.setText(f"Adım {self.mevcut_adim + 1} / 3")

        son = (self.mevcut_adim == 2)
        self.ileri_btn.setVisible(not son)
        self.kaydet_btn.setVisible(son)

        if son:
            self._ozet_doldur()

    def ileri(self):
        if self.mevcut_adim == 0:
            if not self.ad_input.text().strip():
                msg_warn(self, "Eksik Alan", "Lütfen tarif adı girin!")
                return
        self.mevcut_adim += 1
        self._adim_guncelle()

    def geri(self):
        self.mevcut_adim -= 1
        self._adim_guncelle()

    # ── MALZEME İŞLEMLERİ ────────────────────────────────────

    def _malzeme_ekle(self):
        ad = self.w_ad.text().strip()
        miktar = self.w_miktar.text().strip()
        if not ad or not miktar:
            msg_warn(self, "Hata", "Malzeme adı ve miktar giriniz!")
            return
        birim = self.w_birim.currentText()
        fiyat = self.w_fiyat.value()
        kalori = self.w_kalori.value()
        self.malzeme_listesi.append({
            "ad": ad, "miktar": miktar, "birim": birim,
            "fiyat": fiyat, "kalori": kalori
        })
        row = self.w_malzeme_table.rowCount()
        self.w_malzeme_table.insertRow(row)
        self.w_malzeme_table.setItem(row, 0, QTableWidgetItem(ad))
        self.w_malzeme_table.setItem(row, 1, QTableWidgetItem(miktar))
        self.w_malzeme_table.setItem(row, 2, QTableWidgetItem(birim))
        self.w_malzeme_table.setItem(row, 3, QTableWidgetItem(f"₺{fiyat:.2f}"))
        self.w_malzeme_table.setItem(row, 4, QTableWidgetItem(str(kalori)))
        self.w_ad.clear()
        self.w_miktar.clear()
        self.w_fiyat.setValue(0)
        self.w_kalori.setValue(0)
        self._malzeme_ozet_guncelle()

    def _malzeme_sil(self):
        row = self.w_malzeme_table.currentRow()
        if row >= 0:
            self.w_malzeme_table.removeRow(row)
            self.malzeme_listesi.pop(row)
            self._malzeme_ozet_guncelle()

    def _malzeme_ozet_guncelle(self):
        sayi = len(self.malzeme_listesi)
        toplam_maliyet = sum(m["fiyat"] for m in self.malzeme_listesi)
        toplam_kalori = sum(m["kalori"] for m in self.malzeme_listesi)
        self.malzeme_sayisi_lbl.setText(f"{sayi} malzeme")
        self.malzeme_ozet_lbl.setText(
            f"💰 Tahmini Maliyet: ₺{toplam_maliyet:.2f}  |  🔥 Toplam Kalori: {toplam_kalori} cal")

    # ── ÖZET DOLDUR ──────────────────────────────────────────

    def _ozet_doldur(self):
        zorluk_btn = self.zorluk_group.checkedButton()
        zorluk_deger = zorluk_btn.property("deger") if zorluk_btn else 3
        zorluk_map = {1: "👶 Çok Kolay", 2: "😊 Kolay", 3: "😐 Orta", 4: "🤓 Zor", 5: "🔥 Çok Zor"}

        # Grid'i temizle
        while self.ozet_info_layout.count():
            item = self.ozet_info_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        def info_lbl(txt, renk="#f5a623"):
            l = QLabel(txt)
            l.setStyleSheet(f"color: {renk}; font-weight: bold; font-size: 12px;")
            return l

        def val_lbl(txt):
            l = QLabel(txt)
            l.setStyleSheet("color: #ffffff; font-size: 12px;")
            return l

        satirlar = [
            ("📖 Tarif Adı:", self.ad_input.text()),
            ("📚 Kategori:", self.kategori_combo.currentText()),
            ("⏱️ Süre:", f"{self.sure_input.value()} dk"),
            ("🎯 Zorluk:", zorluk_map.get(zorluk_deger, "Orta")),
            ("🥕 Malzeme Sayısı:", f"{len(self.malzeme_listesi)} adet"),
        ]
        for i, (k, v) in enumerate(satirlar):
            self.ozet_info_layout.addWidget(info_lbl(k), i, 0)
            self.ozet_info_layout.addWidget(val_lbl(v), i, 1)

        # Malzeme listesi
        self.ozet_malzeme_list.clear()
        if self.malzeme_listesi:
            toplam_maliyet = sum(m["fiyat"] for m in self.malzeme_listesi)
            toplam_kalori = sum(m["kalori"] for m in self.malzeme_listesi)
            for m in self.malzeme_listesi:
                self.ozet_malzeme_list.addItem(
                    f"  🥕  {m['ad']}  —  {m['miktar']} {m['birim']}  "
                    f"  💰 ₺{m['fiyat']:.2f}  🔥 {m['kalori']} cal")
            self.ozet_besin_lbl.setText(
                f"💰 Toplam Maliyet: ₺{toplam_maliyet:.2f}   "
                f"|   🔥 Toplam Kalori: {toplam_kalori} cal")
        else:
            self.ozet_malzeme_list.addItem("  — Malzeme eklenmedi —")
            self.ozet_besin_lbl.setText("")

    # ── KAYDET ───────────────────────────────────────────────

    def kaydet(self):
        ad = self.ad_input.text().strip()
        if not ad:
            msg_warn(self, "Hata", "Tarif adı girilmedi!")
            return
        zorluk_btn = self.zorluk_group.checkedButton()
        zorluk = zorluk_btn.property("deger") if zorluk_btn else 3
        tarif_id = self.db.tarif_ekle(
            ad,
            self.kategori_combo.currentText(),
            self.sure_input.value(),
            self.talimat_input.toPlainText().strip(),
            zorluk
        )
        for m in self.malzeme_listesi:
            self.db.malzeme_ekle(
                tarif_id, m["ad"], m["miktar"], m["birim"],
                float(m["fiyat"]), int(m["kalori"]), 0.0
            )
        msg_info(self, "Başarılı",
            f"✅ '{ad}' başarıyla eklendi!\n"
            f"📦 {len(self.malzeme_listesi)} malzeme kaydedildi.")
        self.accept()

    def tarifi_kaydet(self):
        """wizard_ac() ile uyumluluk için alias."""
        return True

class TarifEkleDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.secilen_resim = None
        self.setWindowTitle("✨ Yeni Tarif Ekle")
        self.setGeometry(100, 100, 600, 780)
        dialog_maximize_flag(self)
        self.setStyleSheet("""
            QDialog { background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #1a1a2e, stop:1 #2d2d3a); }
            QLabel { color: #f5a623; font-weight: bold; font-size: 12px; }
            QLineEdit, QComboBox, QSpinBox, QTextEdit { background-color: #0f0f1a; color: #ffffff;
                border: 2px solid #f5a623; border-radius: 12px; padding: 12px; font-size: 12px; }
            QLineEdit:focus, QComboBox:focus { border: 2px solid #4CAF50; }
        """)
        self.result = None
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(18)
        layout.setContentsMargins(30, 30, 30, 30)
        baslik = QLabel("✨ YENİ TARİF EKLE")
        baslik.setFont(QFont("Arial", 20, QFont.Bold))
        baslik.setStyleSheet("color: #f5a623;")
        baslik.setAlignment(Qt.AlignCenter)
        layout.addWidget(baslik)

        grid = QGridLayout()
        grid.setSpacing(15)
        grid.addWidget(QLabel("📖 Tarif Adı:"), 0, 0)
        self.ad_input = QLineEdit()
        self.ad_input.setPlaceholderText("Tarif adını girin")
        grid.addWidget(self.ad_input, 0, 1)

        grid.addWidget(QLabel("📚 Kategori:"), 1, 0)
        self.kategori_combo = QComboBox()
        self.kategori_combo.addItems(self.db.kategorileri_getir())
        grid.addWidget(self.kategori_combo, 1, 1)

        grid.addWidget(QLabel("⏱️ Hazırlama Süresi (dk):"), 2, 0)
        self.sure_input = QSpinBox()
        self.sure_input.setRange(1, 600)
        self.sure_input.setValue(30)
        grid.addWidget(self.sure_input, 2, 1)

        grid.addWidget(QLabel("🎯 Zorluk Seviyesi:"), 3, 0)
        self.zorluk_combo = QComboBox()
        self.zorluk_combo.addItems(["👶 Çok Kolay (1)", "😊 Kolay (2)", "😐 Orta (3)", "🤓 Zor (4)", "🔥 Çok Zor (5)"])
        self.zorluk_combo.setCurrentIndex(2)
        grid.addWidget(self.zorluk_combo, 3, 1)

        grid.addWidget(QLabel("🖼️ Resim:"), 4, 0)
        resim_layout = QHBoxLayout()
        self.resim_label = QLabel("Resim seçilmedi")
        self.resim_label.setStyleSheet("color: #999999; font-weight: normal;")
        resim_sec_btn = QPushButton("📁 Seç")
        resim_sec_btn.setStyleSheet("""QPushButton { background-color: #2196F3; color: white; border: none;
            border-radius: 8px; padding: 8px 14px; font-weight: bold; }
            QPushButton:hover { background-color: #1976D2; }""")
        resim_sec_btn.clicked.connect(self.resim_sec)
        resim_layout.addWidget(self.resim_label)
        resim_layout.addWidget(resim_sec_btn)
        grid.addLayout(resim_layout, 4, 1)

        grid.addWidget(QLabel("👨‍🍳 Talimat:"), 5, 0)
        self.talimat_input = QTextEdit()
        self.talimat_input.setMaximumHeight(130)
        self.talimat_input.setPlaceholderText("Tarifin yapılış adımlarını yazın...")
        grid.addWidget(self.talimat_input, 5, 1)
        layout.addLayout(grid)

        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)
        ekle_btn = LuxuryButton("✅ TARİF EKLE", "#4CAF50")
        ekle_btn.clicked.connect(self.ekle)
        iptal_btn = LuxuryButton("❌ İPTAL", "#e63946")
        iptal_btn.clicked.connect(self.reject)
        button_layout.addWidget(ekle_btn)
        button_layout.addWidget(iptal_btn)
        layout.addLayout(button_layout)
        self.setLayout(layout)

    def resim_sec(self):
        yol, _ = QFileDialog.getOpenFileName(self, "Tarif Resmi Seç", "",
                                             "Resim Dosyaları (*.png *.jpg *.jpeg *.bmp *.gif)")
        if yol:
            self.secilen_resim = yol
            self.resim_label.setText(os.path.basename(yol))
            self.resim_label.setStyleSheet("color: #4CAF50; font-weight: bold;")

    def ekle(self):
        ad = self.ad_input.text().strip()
        if not ad:
            msg_warn(self, "Hata", "Tarif adı giriniz!")
            return
        zorluk = self.zorluk_combo.currentIndex() + 1
        self.result = (ad, self.kategori_combo.currentText(), self.sure_input.value(),
                       self.talimat_input.toPlainText().strip(), zorluk, self.secilen_resim)
        self.accept()


class MalzemeEkleDialog(QDialog):
    def __init__(self, db, tarif_id, tarif_adi, parent=None):
        super().__init__(parent)
        self.db = db
        self.tarif_id = tarif_id
        self.setWindowTitle(f"🥕 Malzeme Ekle - {tarif_adi}")
        self.setGeometry(100, 100, 480, 540)
        dialog_maximize_flag(self)
        self.setStyleSheet("""
            QDialog { background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #1a1a2e, stop:1 #2d2d3a); }
            QLabel { color: #4CAF50; font-weight: bold; font-size: 12px; }
            QLineEdit, QSpinBox, QComboBox { background-color: #0f0f1a; color: #ffffff;
                border: 2px solid #4CAF50; border-radius: 12px; padding: 12px; font-size: 12px; }
            QLineEdit:focus { border: 2px solid #f5a623; }
        """)
        self.result = None
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(30, 30, 30, 30)
        baslik = QLabel("🥕 YENİ MALZEME")
        baslik.setFont(QFont("Arial", 18, QFont.Bold))
        baslik.setStyleSheet("color: #4CAF50;")
        baslik.setAlignment(Qt.AlignCenter)
        layout.addWidget(baslik)

        grid = QGridLayout()
        grid.setSpacing(12)
        grid.addWidget(QLabel("Malzeme Adı:"), 0, 0)
        self.malzeme_input = QLineEdit()
        self.malzeme_input.setPlaceholderText("Örn: Domates, Un, Yumurta")
        grid.addWidget(self.malzeme_input, 0, 1)

        grid.addWidget(QLabel("Miktar:"), 1, 0)
        self.miktar_input = QLineEdit()
        self.miktar_input.setPlaceholderText("Örn: 2, 250")
        grid.addWidget(self.miktar_input, 1, 1)

        grid.addWidget(QLabel("Birim:"), 2, 0)
        self.birim_combo = QComboBox()
        self.birim_combo.addItems(["gr", "ml", "adet", "su bardağı", "yemek kaşığı",
                                    "tatlı kaşığı", "çay kaşığı", "kase", "fincan", "diş", "paket"])
        grid.addWidget(self.birim_combo, 2, 1)

        grid.addWidget(QLabel("Fiyat (₺):"), 3, 0)
        self.fiyat_input = QSpinBox()
        self.fiyat_input.setRange(0, 10000)
        grid.addWidget(self.fiyat_input, 3, 1)

        grid.addWidget(QLabel("Kalori:"), 4, 0)
        self.kalori_input = QSpinBox()
        self.kalori_input.setRange(0, 10000)
        grid.addWidget(self.kalori_input, 4, 1)

        grid.addWidget(QLabel("Protein (g):"), 5, 0)
        self.protein_input = QSpinBox()
        self.protein_input.setRange(0, 1000)
        grid.addWidget(self.protein_input, 5, 1)

        layout.addLayout(grid)
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)
        ekle_btn = LuxuryButton("✅ MALZEME EKLE", "#4CAF50")
        ekle_btn.clicked.connect(self.ekle)
        iptal_btn = LuxuryButton("❌ İPTAL", "#e63946")
        iptal_btn.clicked.connect(self.reject)
        button_layout.addWidget(ekle_btn)
        button_layout.addWidget(iptal_btn)
        layout.addLayout(button_layout)
        self.setLayout(layout)

    def ekle(self):
        malzeme = self.malzeme_input.text().strip()
        miktar = self.miktar_input.text().strip()
        if not malzeme or not miktar:
            msg_warn(self, "Hata", "Malzeme adı ve miktar giriniz!")
            return
        self.result = (self.tarif_id, malzeme, miktar, self.birim_combo.currentText(),
                       float(self.fiyat_input.value()), self.kalori_input.value(),
                       float(self.protein_input.value()))
        self.accept()


class KullaniciEkleDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("👤 Yeni Kullanıcı Ekle")
        self.setGeometry(100, 100, 450, 400)
        dialog_maximize_flag(self)
        self.setStyleSheet("""
            QDialog { background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #1a1a2e, stop:1 #2d2d3a); }
            QLabel { color: #f5a623; font-weight: bold; font-size: 12px; }
            QLineEdit { background-color: #0f0f1a; color: #ffffff; border: 2px solid #f5a623;
                border-radius: 12px; padding: 12px; font-size: 12px; }
            QLineEdit:focus { border: 2px solid #4CAF50; }
        """)
        self.result = None
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)
        baslik = QLabel("👤 YENİ KULLANICI")
        baslik.setFont(QFont("Arial", 18, QFont.Bold))
        baslik.setStyleSheet("color: #f5a623;")
        baslik.setAlignment(Qt.AlignCenter)
        layout.addWidget(baslik)
        grid = QGridLayout()
        grid.setSpacing(15)
        grid.addWidget(QLabel("Ad:"), 0, 0)
        self.ad_input = QLineEdit()
        grid.addWidget(self.ad_input, 0, 1)
        grid.addWidget(QLabel("Soyad:"), 1, 0)
        self.soyad_input = QLineEdit()
        grid.addWidget(self.soyad_input, 1, 1)
        grid.addWidget(QLabel("Email:"), 2, 0)
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("ornek@email.com")
        grid.addWidget(self.email_input, 2, 1)
        layout.addLayout(grid)
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)
        ekle_btn = LuxuryButton("✅ KULLANICI EKLE", "#4CAF50")
        ekle_btn.clicked.connect(self.ekle)
        iptal_btn = LuxuryButton("❌ İPTAL", "#e63946")
        iptal_btn.clicked.connect(self.reject)
        button_layout.addWidget(ekle_btn)
        button_layout.addWidget(iptal_btn)
        layout.addLayout(button_layout)
        self.setLayout(layout)

    def ekle(self):
        ad = self.ad_input.text().strip()
        soyad = self.soyad_input.text().strip()
        email = self.email_input.text().strip()
        if not ad or not soyad:
            msg_warn(self, "Hata", "Ad ve Soyad zorunludur!")
            return
        if email and "@" not in email:
            msg_warn(self, "Hata", "Geçerli bir email giriniz!")
            return
        self.result = (ad, soyad, email if email else None)
        self.accept()


class DegerlendirmeDialog(QDialog):
    def __init__(self, db, tarif_id, tarif_adi, parent=None):
        super().__init__(parent)
        self.db = db
        self.tarif_id = tarif_id
        self.setWindowTitle(f"⭐ Değerlendir - {tarif_adi}")
        self.setGeometry(100, 100, 480, 420)
        dialog_maximize_flag(self)
        self.setStyleSheet("""
            QDialog { background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #1a1a2e, stop:1 #2d2d3a); }
            QLabel { color: #f5a623; font-weight: bold; font-size: 12px; }
            QComboBox, QTextEdit { background-color: #0f0f1a; color: #ffffff;
                border: 2px solid #f5a623; border-radius: 12px; padding: 12px; font-size: 12px; }
            QComboBox:focus { border: 2px solid #4CAF50; }
        """)
        self.result = None
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(18)
        layout.setContentsMargins(30, 30, 30, 30)

        baslik = QLabel("⭐ TARİFİ DEĞERLENDİR")
        baslik.setFont(QFont("Arial", 18, QFont.Bold))
        baslik.setStyleSheet("color: #f5a623;")
        baslik.setAlignment(Qt.AlignCenter)
        layout.addWidget(baslik)

        grid = QGridLayout()
        grid.setSpacing(15)

        grid.addWidget(QLabel("Kullanıcı:"), 0, 0)
        self.kullanici_combo = QComboBox()
        kullanicilar = self.db.kullanicilari_getir()
        if kullanicilar:
            for k in kullanicilar:
                self.kullanici_combo.addItem(f"{k['ad']} {k['soyad']}", k['kullanici_id'])
        else:
            self.kullanici_combo.addItem("— Önce kullanıcı ekleyin —", None)
        grid.addWidget(self.kullanici_combo, 0, 1)

        grid.addWidget(QLabel("Puan (1-5):"), 1, 0)
        self.puan_combo = QComboBox()
        self.puan_combo.addItems(["⭐⭐⭐⭐⭐ (5)", "⭐⭐⭐⭐ (4)", "⭐⭐⭐ (3)", "⭐⭐ (2)", "⭐ (1)"])
        grid.addWidget(self.puan_combo, 1, 1)

        grid.addWidget(QLabel("Yorum:"), 2, 0)
        self.yorum_input = QTextEdit()
        self.yorum_input.setMaximumHeight(110)
        self.yorum_input.setPlaceholderText("Tarif hakkında yorumunuz...")
        grid.addWidget(self.yorum_input, 2, 1)

        layout.addLayout(grid)

        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)
        kaydet_btn = LuxuryButton("✅ DEĞERLENDİR", "#4CAF50")
        kaydet_btn.clicked.connect(self.kaydet)
        iptal_btn = LuxuryButton("❌ İPTAL", "#e63946")
        iptal_btn.clicked.connect(self.reject)
        button_layout.addWidget(kaydet_btn)
        button_layout.addWidget(iptal_btn)
        layout.addLayout(button_layout)
        self.setLayout(layout)

    def kaydet(self):
        kullanici_id = self.kullanici_combo.currentData()
        if not kullanici_id:
            msg_warn(self, "Hata", "Önce Kullanıcılar sekmesinden kullanıcı ekleyin!")
            return
        puan_text = self.puan_combo.currentText()
        puan = int(puan_text.split("(")[1].split(")")[0])
        self.result = (self.tarif_id, kullanici_id, puan, self.yorum_input.toPlainText().strip())
        self.accept()


# ===================== TIER 1: GELİŞMİŞ ARAMA DİALOGU =====================

class AramaDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("🔍 Gelişmiş Arama")
        self.setGeometry(150, 150, 550, 520)
        dialog_maximize_flag(self)
        self.setStyleSheet("""
            QDialog { background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #1a1a2e, stop:1 #2d2d3a); }
            QLabel { color: #f5a623; font-weight: bold; font-size: 12px; }
            QLineEdit, QComboBox, QSpinBox { background-color: #0f0f1a; color: #ffffff;
                border: 2px solid #f5a623; border-radius: 10px; padding: 10px; font-size: 12px; }
            QLineEdit:focus, QComboBox:focus { border: 2px solid #4CAF50; }
            QCheckBox { color: #f5a623; font-weight: bold; font-size: 12px; }
        """)
        self.results = None
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(18)
        layout.setContentsMargins(30, 30, 30, 30)
        title = QLabel("🔍 GELİŞMİŞ ARAMA VE FİLTRELEME")
        title.setFont(QFont("Arial", 16, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        grid = QGridLayout()
        grid.setSpacing(15)
        grid.addWidget(QLabel("📖 Tarif / Kategori / Malzeme:"), 0, 0)
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Örn: Mantı, Domates, Pide...")
        grid.addWidget(self.search_input, 0, 1)
        grid.addWidget(QLabel("⭐ Zorluk Seviyesi:"), 1, 0)
        self.zorluk_combo = QComboBox()
        self.zorluk_combo.addItems(["Tümü", "👶 Çok Kolay (1)", "😊 Kolay (2)",
                                    "😐 Orta (3)", "🤓 Zor (4)", "🔥 Çok Zor (5)"])
        grid.addWidget(self.zorluk_combo, 1, 1)
        grid.addWidget(QLabel("⏱️ Süre (dk) - Min / Max:"), 2, 0)
        sure_layout = QHBoxLayout()
        self.min_sure = QSpinBox()
        self.min_sure.setRange(0, 600)
        sure_layout.addWidget(self.min_sure)
        self.max_sure = QSpinBox()
        self.max_sure.setRange(0, 600)
        self.max_sure.setValue(600)
        sure_layout.addWidget(self.max_sure)
        grid.addLayout(sure_layout, 2, 1)
        grid.addWidget(QLabel("💰 Maliyet (₺) - Min / Max:"), 3, 0)
        maliyet_layout = QHBoxLayout()
        self.min_maliyet = QSpinBox()
        self.min_maliyet.setRange(0, 5000)
        maliyet_layout.addWidget(self.min_maliyet)
        self.max_maliyet = QSpinBox()
        self.max_maliyet.setRange(0, 5000)
        self.max_maliyet.setValue(5000)
        maliyet_layout.addWidget(self.max_maliyet)
        grid.addLayout(maliyet_layout, 3, 1)
        layout.addLayout(grid)
        self.favorit_check = QCheckBox("⭐ Sadece Favorileri Göster")
        layout.addWidget(self.favorit_check)
        layout.addStretch()
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(15)
        ara_btn = LuxuryButton("🔍 ARA", "#4CAF50")
        ara_btn.clicked.connect(self.ara)
        temizle_btn = LuxuryButton("🔄 TEMİZLE", "#2196F3")
        temizle_btn.clicked.connect(self.temizle)
        iptal_btn = LuxuryButton("❌ İPTAL", "#e63946")
        iptal_btn.clicked.connect(self.reject)
        btn_layout.addWidget(ara_btn)
        btn_layout.addWidget(temizle_btn)
        btn_layout.addWidget(iptal_btn)
        layout.addLayout(btn_layout)
        self.setLayout(layout)

    def ara(self):
        arama_terimi = self.search_input.text().strip()
        sonuclar = self.db.tarif_ara(arama_terimi) if arama_terimi else self.db.tarifleri_getir()
        sonuclar = [t for t in sonuclar
                    if self.min_sure.value() <= t['hazirlama_suresi'] <= self.max_sure.value()]
        zorluk_text = self.zorluk_combo.currentText()
        if zorluk_text != "Tümü":
            zorluk = int(zorluk_text.split("(")[1].split(")")[0])
            sonuclar = [t for t in sonuclar if t['zorluk_seviyesi'] == zorluk]
        filtered = []
        for t in sonuclar:
            maliyet = self.db.tahmini_tarif_maliyeti(t['tarif_id'])
            if self.min_maliyet.value() <= maliyet <= self.max_maliyet.value():
                filtered.append(t)
        sonuclar = filtered
        if self.favorit_check.isChecked():
            sonuclar = [t for t in sonuclar if t['is_favorite']]
        self.results = sonuclar
        msg_info(self, "Sonuç", f"✅ {len(sonuclar)} tarif bulundu!") if sonuclar else \
            msg_warn(self, "Sonuç Yok", "Arama kriterlerine uygun tarif bulunamadı.")
        self.accept()

    def temizle(self):
        self.search_input.clear()
        self.zorluk_combo.setCurrentIndex(0)
        self.min_sure.setValue(0)
        self.max_sure.setValue(600)
        self.min_maliyet.setValue(0)
        self.max_maliyet.setValue(5000)
        self.favorit_check.setChecked(False)


# ===================== TARİF DETAY DİALOGU (TIER 1/2/3/4) =====================

class TarifDetayDialog(QDialog):
    def __init__(self, db, tarif_id, parent=None):
        super().__init__(parent)
        self.db = db
        self.tarif_id = tarif_id
        self.tarif = db.tarif_bul(tarif_id)
        self.malzemeler = db.malzemeleri_getir(tarif_id)
        self.degerlendirmeler = db.degerlendirmeleri_getir(tarif_id)
        self.setWindowTitle(f"🍽️ {self.tarif['ad']}")
        self.setGeometry(100, 100, 1050, 860)
        dialog_maximize_flag(self)
        self.setStyleSheet("""
            QDialog { background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #1a1a2e, stop:1 #2d2d3a); }
            QLabel { color: #ffffff; }
            QTabBar::tab { background-color: #1a1a2e; color: #f5a623; padding: 12px 28px;
                border-radius: 8px; margin-right: 5px; border: 1px solid #f5a623; }
            QTabBar::tab:selected { background-color: #f5a623; color: #1a1a2e; }
            QTableWidget { background-color: #2d2d3a; alternate-background-color: #3d3d4a;
                color: #ffffff; gridline-color: #4CAF50; border: 1px solid #f5a623; border-radius: 10px; }
            QTableWidget::item { padding: 8px; }
            QHeaderView::section { background-color: #1a1a2e; color: #f5a623; font-weight: bold; padding: 10px; border: none; }
            QTextEdit { background-color: #0f0f1a; color: #ffffff; border: 1px solid #f5a623;
                border-radius: 10px; padding: 20px; font-size: 13px; }
            QListWidget { background-color: #0f0f1a; color: #ffffff; border: 1px solid #4CAF50; border-radius: 10px; }
        """)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.addWidget(self.create_header())

        tabs = QTabWidget()
        tabs.addTab(self.create_malzeme_tab(), "🥕 Malzemeler")
        tabs.addTab(self.create_talimat_tab(), "👨‍🍳 Talimat")
        tabs.addTab(self.create_degerlendirme_tab(), "⭐ Değerlendirmeler")
        tabs.addTab(self.create_notlar_tab(), "📝 Notlarım")        # TIER 4
        tabs.addTab(self.create_versiyon_tab(), "📜 Versiyonlar")   # TIER 4
        layout.addWidget(tabs)

        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        favorit_text = "⭐ FAVORİDEN ÇIKAR" if self.tarif['is_favorite'] else "⭐ FAVORİYE EKLE"
        self.favorit_btn = LuxuryButton(favorit_text, "#f5a623")
        self.favorit_btn.clicked.connect(self.toggle_favorite)
        yaptim_btn = LuxuryButton("✅ YAPTIM", "#4CAF50")
        yaptim_btn.clicked.connect(self.yaptim_kaydet)
        resim_btn = LuxuryButton("🖼️ RESİM", "#9c27b0")
        resim_btn.clicked.connect(self.resim_degistir)
        not_ekle_btn = LuxuryButton("📝 NOT EKLE", "#2196F3")  # TIER 4
        not_ekle_btn.clicked.connect(self.not_ekle_tarife)
        zorluk_btn = LuxuryButton("🎯 ZORLUK", "#795548")
        zorluk_btn.clicked.connect(self.set_zorluk)
        kapat_btn = LuxuryButton("❌ KAPAT", "#e63946")
        kapat_btn.clicked.connect(self.close)
        btn_layout.addWidget(self.favorit_btn)
        btn_layout.addWidget(yaptim_btn)
        btn_layout.addWidget(resim_btn)
        btn_layout.addWidget(not_ekle_btn)
        btn_layout.addWidget(zorluk_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(kapat_btn)
        layout.addLayout(btn_layout)
        self.setLayout(layout)

    def get_zorluk_etiketi(self, zorluk):
        return {1: "👶 Çok Kolay", 2: "😊 Kolay", 3: "😐 Orta", 4: "🤓 Zor", 5: "🔥 Çok Zor"}.get(zorluk, "?")

    def create_header(self):
        header = QFrame()
        header.setStyleSheet("""
            QFrame { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 15px; border: 1.5px solid #f5a623; }
        """)
        header.setFixedHeight(180)
        ana = QHBoxLayout()
        ana.setContentsMargins(20, 15, 20, 15)
        ana.setSpacing(20)

        self.resim_preview = QLabel()
        self.resim_preview.setFixedSize(150, 150)
        self.resim_preview.setStyleSheet("border: 1px solid #f5a623; border-radius: 10px; background-color: #0f0f1a;")
        self.resim_preview.setAlignment(Qt.AlignCenter)
        self._resim_yukle()
        ana.addWidget(self.resim_preview)

        info_layout = QVBoxLayout()
        info_layout.setSpacing(6)
        title = QLabel(self.tarif['ad'])
        title.setFont(QFont("Arial", 20, QFont.Bold))
        title.setStyleSheet("color: #f5a623;")
        info_layout.addWidget(title)

        puan = self.db.ortalama_puan(self.tarif_id)
        zorluk = self.get_zorluk_etiketi(self.tarif['zorluk_seviyesi'])
        info1 = QLabel(f"📚 {self.tarif['kategori']}  |  ⏱️ {self.tarif['hazirlama_suresi']} dk  |  🎯 {zorluk}")
        info1.setStyleSheet("color: #ffffff; font-size: 12px;")
        info_layout.addWidget(info1)

        olusturma = str(self.tarif.get('olusturma_tarihi', '') or '')
        info2 = QLabel(f"⭐ Puan: {puan}/5  |  📅 Eklenme: {olusturma[:10]}")
        info2.setStyleSheet("color: #4CAF50; font-size: 12px;")
        info_layout.addWidget(info2)

        maliyet = self.db.tahmini_tarif_maliyeti(self.tarif_id)
        kalori = self.db.tahmini_tarif_kalorileri(self.tarif_id)
        protein = self.db.tahmini_tarif_proteini(self.tarif_id)
        stats = QLabel(f"💰 ₺{maliyet:.2f}  |  🔥 {kalori} cal  |  💪 {protein:.1f}g")
        stats.setStyleSheet("color: #f5a623; font-size: 12px; font-weight: bold;")
        info_layout.addWidget(stats)
        info_layout.addStretch()
        ana.addLayout(info_layout)
        header.setLayout(ana)
        return header

    def _resim_yukle(self):
        yol = self.tarif.get('goruntu_yolu')
        if yol and os.path.exists(yol):
            pixmap = QPixmap(yol)
            if not pixmap.isNull():
                self.resim_preview.setPixmap(pixmap.scaled(150, 150, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                return
        self.resim_preview.setText("🍽️\nResim yok")
        self.resim_preview.setStyleSheet("border: 1px solid #f5a623; border-radius: 10px; background-color: #0f0f1a; color: #999999;")

    def resim_degistir(self):
        yol, _ = QFileDialog.getOpenFileName(self, "Tarif Resmi Seç", "",
                                             "Resim Dosyaları (*.png *.jpg *.jpeg *.bmp *.gif)")
        if yol:
            self.db.tarif_resim_guncelle(self.tarif_id, yol)
            self.tarif['goruntu_yolu'] = yol
            self._resim_yukle()
            msg_info(self, "Başarılı", "✅ Resim güncellendi!")

    def create_malzeme_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels(["Malzeme", "Miktar", "Birim", "Fiyat (₺)", "Kalori", "Protein (g)"])
        table.setAlternatingRowColors(True)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        toplam_fiyat = toplam_kalori = toplam_protein = 0.0
        for m in self.malzemeler:
            row = table.rowCount()
            table.insertRow(row)
            table.setItem(row, 0, QTableWidgetItem(m['malzeme_adi']))
            table.setItem(row, 1, QTableWidgetItem(str(m['miktar'])))
            table.setItem(row, 2, QTableWidgetItem(m.get('birim', 'gr')))
            fiyat = m.get('fiyat', 0) or 0
            table.setItem(row, 3, QTableWidgetItem(f"₺{fiyat:.2f}"))
            kalori = m.get('kalori', 0) or 0
            table.setItem(row, 4, QTableWidgetItem(str(kalori)))
            protein = m.get('protein', 0) or 0
            table.setItem(row, 5, QTableWidgetItem(f"{protein:.1f}"))
            toplam_fiyat += fiyat
            toplam_kalori += kalori
            toplam_protein += protein
        layout.addWidget(table)

        ozet_frame = QFrame()
        ozet_frame.setStyleSheet("QFrame { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #2d2d3a, stop:1 #1a1a2e); border-radius: 12px; border: 1px solid #4CAF50; padding: 15px; }")
        ozet_layout = QHBoxLayout()
        for baslik, deger, renk in [
            ("💰 Toplam Maliyet", f"₺{toplam_fiyat:.2f}", "#4CAF50"),
            ("🔥 Toplam Kalori", f"{int(toplam_kalori)} cal", "#f5a623"),
            ("💪 Toplam Protein", f"{toplam_protein:.1f}g", "#2196F3"),
        ]:
            box = QVBoxLayout()
            t = QLabel(baslik)
            t.setStyleSheet(f"color: {renk}; font-weight: bold; font-size: 11px;")
            v = QLabel(deger)
            v.setStyleSheet(f"color: {renk}; font-size: 16px; font-weight: bold;")
            box.addWidget(t)
            box.addWidget(v)
            ozet_layout.addLayout(box)
        ozet_layout.addStretch()
        ozet_frame.setLayout(ozet_layout)
        layout.addWidget(ozet_frame)
        widget.setLayout(layout)
        return widget

    def create_talimat_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        talimat = QTextEdit()
        talimat.setText(self.tarif['talimat'] or "Talimat bulunmuyor")
        talimat.setReadOnly(True)
        layout.addWidget(talimat)
        widget.setLayout(layout)
        return widget

    def create_degerlendirme_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(12)
        if not self.degerlendirmeler:
            bos = QLabel("Henüz değerlendirme yapılmamış")
            bos.setAlignment(Qt.AlignCenter)
            bos.setStyleSheet("color: #999999; font-style: italic; padding: 40px;")
            layout.addWidget(bos)
        else:
            for d in self.degerlendirmeler:
                card = QFrame()
                card.setStyleSheet("QFrame { background-color: #2d2d3a; border-radius: 10px; padding: 15px; border-left: 4px solid #f5a623; }")
                cl = QVBoxLayout()
                cl.setSpacing(6)
                k = QLabel(f"👤 {d['ad']} {d['soyad']}")
                k.setStyleSheet("color: #f5a623; font-weight: bold; font-size: 12px;")
                p = QLabel("⭐ " * d['puan'])
                y = QLabel(d['yorum'] or "(Yorum yapılmadı)")
                y.setStyleSheet("color: #ffffff; font-style: italic; font-size: 12px;")
                y.setWordWrap(True)
                t = QLabel(d['tarih'][:10])
                t.setStyleSheet("color: #999999; font-size: 10px;")
                cl.addWidget(k); cl.addWidget(p); cl.addWidget(y); cl.addWidget(t)
                card.setLayout(cl)
                layout.addWidget(card)
        layout.addStretch()
        widget.setLayout(layout)
        return widget

    # TIER 4: Notlar sekmesi
    def create_notlar_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(10)

        notlar = self.db.notlari_getir(self.tarif_id)
        if not notlar:
            bos = QLabel("Bu tarif için henüz not yok.\nAşağıdaki butonu kullanarak not ekleyebilirsiniz.")
            bos.setAlignment(Qt.AlignCenter)
            bos.setStyleSheet("color: #999999; font-style: italic; padding: 30px;")
            layout.addWidget(bos)
        else:
            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
            scroll_widget = QWidget()
            scroll_widget.setStyleSheet("background: transparent;")
            scroll_layout = QVBoxLayout(scroll_widget)
            scroll_layout.setSpacing(10)
            for n in notlar:
                renk = n.get('renk', '#f5a623')
                card = QFrame()
                card.setStyleSheet(f"QFrame {{ background-color: #2d2d3a; border-radius: 10px; border-left: 5px solid {renk}; padding: 10px; }}")
                cl = QVBoxLayout()
                title = QLabel(f"📌 {n['baslik']}")
                title.setStyleSheet(f"color: {renk}; font-weight: bold;")
                cl.addWidget(title)
                if n.get('icerik'):
                    ic = QLabel(n['icerik'][:150] + ("..." if len(n['icerik']) > 150 else ""))
                    ic.setStyleSheet("color: #cccccc; font-size: 11px;")
                    ic.setWordWrap(True)
                    cl.addWidget(ic)
                tarih = QLabel(f"🕐 {str(n['guncelleme_tarihi'])[:16]}")
                tarih.setStyleSheet("color: #777777; font-size: 10px;")
                cl.addWidget(tarih)
                card.setLayout(cl)
                scroll_layout.addWidget(card)
            scroll_layout.addStretch()
            scroll.setWidget(scroll_widget)
            layout.addWidget(scroll)

        ekle_btn = LuxuryButton("📝 BU TARİFE NOT EKLE", "#2196F3")
        ekle_btn.clicked.connect(self.not_ekle_tarife)
        layout.addWidget(ekle_btn)
        widget.setLayout(layout)
        return widget

    # TIER 4: Versiyonlar sekmesi
    def create_versiyon_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(10)
        versiyonlar = self.db.versiyonlari_getir(self.tarif_id)
        if not versiyonlar:
            bos = QLabel("Henüz versiyon geçmişi yok.\nTarifı düzenlediğinizde otomatik snapshot alınır.")
            bos.setAlignment(Qt.AlignCenter)
            bos.setStyleSheet("color: #999999; font-style: italic; padding: 30px;")
            layout.addWidget(bos)
        else:
            table = QTableWidget()
            table.setColumnCount(4)
            table.setHorizontalHeaderLabels(["Versiyon", "Tarih", "Değişiklik Notu", "Ad"])
            table.setAlternatingRowColors(True)
            table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            for v in versiyonlar:
                row = table.rowCount()
                table.insertRow(row)
                table.setItem(row, 0, QTableWidgetItem(f"v{v['versiyon_no']}"))
                table.setItem(row, 1, QTableWidgetItem(str(v['kayit_tarihi'])[:16]))
                table.setItem(row, 2, QTableWidgetItem(v['degisiklik_notu'] or '-'))
                table.setItem(row, 3, QTableWidgetItem(v['ad']))
            layout.addWidget(table)

        goster_btn = LuxuryButton("📜 TAM VERSİYON GEÇMİŞİ", "#9c27b0")
        goster_btn.clicked.connect(self.versiyon_gecmis_ac)
        layout.addWidget(goster_btn)
        widget.setLayout(layout)
        return widget

    def not_ekle_tarife(self):
        dialog = KisiselNotDialog(self.db, tarif_id=self.tarif_id, parent=self)
        if dialog.exec_() == QDialog.Accepted and dialog.result:
            self.db.not_ekle(dialog.result['baslik'], dialog.result['icerik'],
                             dialog.result['renk'], self.tarif_id)
            msg_info(self, "Başarılı", "✅ Not eklendi!")

    def versiyon_gecmis_ac(self):
        dialog = VersiyonGecmisDialog(self.db, self.tarif_id, self)
        dialog.exec_()

    def set_zorluk(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("🎯 Zorluk Seviyesi Ayarla")
        dialog.setGeometry(200, 200, 420, 240)
        dialog.setStyleSheet("""
            QDialog { background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                stop:0 #1a1a2e, stop:1 #2d2d3a); }
            QLabel { color: #f5a623; font-weight: bold; }
            QComboBox { background-color: #0f0f1a; color: #ffffff;
                border: 2px solid #f5a623; border-radius: 10px; padding: 10px; }
        """)
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(25, 25, 25, 25)
        layout.addWidget(QLabel("🎯 Zorluk Seviyesi Seç:"))
        combo = QComboBox()
        combo.addItems(["👶 Çok Kolay (1)", "😊 Kolay (2)", "😐 Orta (3)",
                        "🤓 Zor (4)", "🔥 Çok Zor (5)"])
        mevcut = self.tarif.get('zorluk_seviyesi') or 3
        combo.setCurrentIndex(mevcut - 1)
        layout.addWidget(combo)
        layout.addStretch()
        def kaydet():
            yeni = combo.currentIndex() + 1
            self.db.tarif_zorluk_guncelle(self.tarif_id, yeni)
            self.tarif['zorluk_seviyesi'] = yeni
            msg_info(self, "Başarılı", "✅ Zorluk güncellendi!")
            dialog.close()
        kaydet_btn = LuxuryButton("✅ KAYDET", "#4CAF50")
        kaydet_btn.clicked.connect(kaydet)
        iptal_btn = LuxuryButton("❌ İPTAL", "#e63946")
        iptal_btn.clicked.connect(dialog.close)
        bl = QHBoxLayout()
        bl.addWidget(kaydet_btn)
        bl.addWidget(iptal_btn)
        layout.addLayout(bl)
        dialog.setLayout(layout)
        dialog.exec_()

    def toggle_favorite(self):
        is_favorite = not self.tarif['is_favorite']
        self.db.tarif_favoriye_ekle(self.tarif_id, is_favorite)
        self.tarif['is_favorite'] = 1 if is_favorite else 0
        msg_info(self, "Başarılı",
                               "✅ Favorilere eklendi!" if is_favorite else "✅ Favorilerden çıkarıldı!")
        self.favorit_btn.setText("⭐ FAVORİDEN ÇIKAR" if is_favorite else "⭐ FAVORİYE EKLE")

    def yaptim_kaydet(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("✅ Bu Tarifi Yaptım")
        dialog.setGeometry(200, 200, 420, 280)
        dialog.setStyleSheet("QDialog { background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #1a1a2e, stop:1 #2d2d3a); } QLabel { color: #4CAF50; font-weight: bold; } QTextEdit { background-color: #0f0f1a; color: #ffffff; border: 2px solid #4CAF50; border-radius: 10px; padding: 10px; }")
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(25, 25, 25, 25)
        layout.addWidget(QLabel("Notunuz (opsiyonel):"))
        not_input = QTextEdit()
        not_input.setPlaceholderText("Örn: Soğanı az ekledim, çok lezzetli oldu...")
        layout.addWidget(not_input)
        def kaydet():
            self.db.yapilmis_log_ekle(self.tarif_id, not_input.toPlainText().strip())
            msg_info(self, "Başarılı", "✅ Tarifin yapıldığı kaydedildi!")
            dialog.close()
        kaydet_btn = LuxuryButton("✅ KAYDET", "#4CAF50")
        kaydet_btn.clicked.connect(kaydet)
        iptal_btn = LuxuryButton("❌ İPTAL", "#e63946")
        iptal_btn.clicked.connect(dialog.close)
        bl = QHBoxLayout()
        bl.addWidget(kaydet_btn)
        bl.addWidget(iptal_btn)
        layout.addLayout(bl)
        dialog.setLayout(layout)
        dialog.exec_()


# ===================== GRAFİKLER (TIER 3: 6 grafik) =====================

class StatisticsWidget(QWidget):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)

        # Üst özet satırı
        ozet_layout = QHBoxLayout()
        ozet_layout.setSpacing(10)
        self.ozet_labels = {}
        ozet_items = [
            ("toplam_tarif", "📖 Tarif", "#f5a623"),
            ("toplam_kullanici", "👥 Kullanıcı", "#2196F3"),
            ("toplam_degerlendirme", "⭐ Değerl.", "#e63946"),
            ("favoriler_sayi", "❤️ Favori", "#9c27b0"),
            ("toplam_yapilmis", "✅ Yapılan", "#4CAF50"),
            ("toplam_not", "📝 Not", "#795548"),
            ("toplam_versiyon", "📜 Versiyon", "#00BCD4"),
        ]
        for key, etiket, renk in ozet_items:
            kart = QFrame()
            kart.setStyleSheet(f"""
                QFrame {{
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                        stop:0 #2d2d3a, stop:1 #1a1a2e);
                    border-radius: 10px; border-left: 4px solid {renk};
                    padding: 5px;
                }}
            """)
            kl = QVBoxLayout(kart)
            kl.setSpacing(2)
            kl.setContentsMargins(10, 8, 10, 8)
            et = QLabel(etiket)
            et.setStyleSheet(f"color: {renk}; font-size: 10px; font-weight: bold;")
            val = QLabel("0")
            val.setStyleSheet("color: #ffffff; font-size: 18px; font-weight: bold;")
            kl.addWidget(et)
            kl.addWidget(val)
            ozet_layout.addWidget(kart)
            self.ozet_labels[key] = val
        layout.addLayout(ozet_layout)

        # Yenile butonu
        yenile_btn = LuxuryButton("🔄 GRAFİKLERİ YENİLE", "#4CAF50")
        yenile_btn.setMaximumHeight(36)
        yenile_btn.setMaximumWidth(220)
        yenile_btn.clicked.connect(self.update_charts)
        layout.addWidget(yenile_btn)

        self.figure = Figure(figsize=(14, 7), dpi=100, facecolor='#1a1a2e')
        self.canvas = FigureCanvas(self.figure)
        self.canvas.setStyleSheet("background-color: #1a1a2e; border-radius: 20px;")
        layout.addWidget(self.canvas)
        self.setLayout(layout)

    def _guncelle_ozet(self, ist):
        for key, lbl in self.ozet_labels.items():
            lbl.setText(str(ist.get(key, 0)))

    def _zorluk_etiketi(self, z):
        return {1: "Çok Kolay", 2: "Kolay", 3: "Orta", 4: "Zor", 5: "Çok Zor"}.get(z, str(z))

    def update_charts(self):
        self.figure.clear()
        ist = self.db.istatistikler()
        self._guncelle_ozet(ist)
        renkler = ['#f5a623', '#4CAF50', '#2196F3', '#e63946', '#9c27b0', '#00BCD4', '#795548', '#E91E63', '#607D8B']

        ax1 = self.figure.add_subplot(231)
        ax1.set_facecolor('#2d2d3a')
        if ist['kategori_dagilimi']:
            kategoriler = [k['kategori'][:10] for k in ist['kategori_dagilimi']]
            sayilar = [k['sayi'] for k in ist['kategori_dagilimi']]
            w, texts, autotexts = ax1.pie(sayilar, labels=kategoriler, autopct='%1.0f%%',
                                          colors=renkler[:len(kategoriler)], startangle=90)
            for tx in texts:
                tx.set_color('#ffffff'); tx.set_fontsize(7)
            for tx in autotexts:
                tx.set_color('#ffffff'); tx.set_fontweight('bold'); tx.set_fontsize(7)
        ax1.set_title('📊 Kategori Dağılımı', fontsize=10, fontweight='bold', color='#f5a623')

        ax2 = self.figure.add_subplot(232)
        ax2.set_facecolor('#2d2d3a')
        # TIER 4: 5 bar (+ Not + Versiyon)
        labels = ['Tarif', 'Kategori', 'Kullanıcı', 'Not', 'Versiyon']
        values = [ist['toplam_tarif'], ist['toplam_kategori'], ist['toplam_kullanici'],
                  ist['toplam_not'], ist['toplam_versiyon']]
        bars = ax2.bar(labels, values, color=['#f5a623', '#4CAF50', '#2196F3', '#9c27b0', '#00BCD4'], width=0.6)
        ax2.set_title('📈 Sistem İstatistikleri', fontsize=10, fontweight='bold', color='#f5a623')
        ax2.tick_params(axis='y', colors='#ffffff', labelsize=8)
        ax2.tick_params(axis='x', colors='#ffffff', labelsize=8)
        for b in bars:
            h = b.get_height()
            ax2.text(b.get_x() + b.get_width()/2., h + 0.2, f'{int(h)}',
                     ha='center', va='bottom', fontweight='bold', fontsize=8, color='#f5a623')

        ax3 = self.figure.add_subplot(233)
        ax3.set_facecolor('#2d2d3a')
        zd = self.db.zorluk_dagilimi()
        if zd:
            zl = [self._zorluk_etiketi(z['zorluk_seviyesi']) for z in zd]
            zs = [z['sayi'] for z in zd]
            ax3.bar(zl, zs, color='#9c27b0', width=0.6)
        ax3.set_title('🎯 Zorluk Dağılımı', fontsize=10, fontweight='bold', color='#f5a623')
        ax3.tick_params(axis='y', colors='#ffffff', labelsize=8)
        ax3.tick_params(axis='x', colors='#ffffff', labelsize=7, rotation=20)

        ax4 = self.figure.add_subplot(234)
        ax4.set_facecolor('#2d2d3a')
        ecy = self.db.en_cok_yapilan_tarifler(5)
        if ecy:
            adlar = [t['ad'][:14] for t in ecy][::-1]
            sayilar_ecy = [t['sayi'] for t in ecy][::-1]
            ax4.barh(adlar, sayilar_ecy, color='#4CAF50')
        else:
            ax4.text(0.5, 0.5, 'Henüz "Yaptım" kaydı yok', ha='center', va='center',
                     color='#999999', fontsize=9, transform=ax4.transAxes)
        ax4.set_title('🏆 En Çok Yapılanlar', fontsize=10, fontweight='bold', color='#f5a623')
        ax4.tick_params(axis='y', colors='#ffffff', labelsize=7)
        ax4.tick_params(axis='x', colors='#ffffff', labelsize=8)

        ax5 = self.figure.add_subplot(235)
        ax5.set_facecolor('#2d2d3a')
        pd_ = self.db.puan_dagilimi()
        if pd_:
            puanlar = [f"{p['puan']}⭐" for p in pd_]
            sayilar_p = [p['sayi'] for p in pd_]
            ax5.bar(puanlar, sayilar_p, color='#e63946', width=0.6)
        else:
            ax5.text(0.5, 0.5, 'Henüz değerlendirme yok', ha='center', va='center',
                     color='#999999', fontsize=9, transform=ax5.transAxes)
        ax5.set_title('⭐ Puan Dağılımı', fontsize=10, fontweight='bold', color='#f5a623')
        ax5.tick_params(axis='y', colors='#ffffff', labelsize=8)
        ax5.tick_params(axis='x', colors='#ffffff', labelsize=8)

        ax6 = self.figure.add_subplot(236)
        ax6.set_facecolor('#2d2d3a')
        ay = self.db.aylik_yapilmis_sayisi()
        if ay:
            aylar = [a['ay'] for a in ay]
            sayilar_ay = [a['sayi'] for a in ay]
            ax6.plot(aylar, sayilar_ay, marker='o', color='#2196F3', linewidth=2)
        else:
            ax6.text(0.5, 0.5, 'Henüz "Yaptım" kaydı yok', ha='center', va='center',
                     color='#999999', fontsize=9, transform=ax6.transAxes)
        ax6.set_title('📅 Aylık Yapılma Trendi', fontsize=10, fontweight='bold', color='#f5a623')
        ax6.tick_params(axis='y', colors='#ffffff', labelsize=8)
        ax6.tick_params(axis='x', colors='#ffffff', labelsize=7, rotation=30)

        self.figure.tight_layout()
        self.canvas.draw()


# ===================== ANA PENCERE =====================

class YemekTarifPlatformu(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = DatabaseManager()
        self.current_theme = "dark"
        self.setWindowTitle("🍽️ YEMEK TARİF PLATFORMU | LUXURY EDITION")
        self.setGeometry(50, 50, 1480, 920)
        self.setStyleSheet(ThemeManager.get(self.current_theme))
        self.init_ui()
        self.setup_shortcuts()
        self.load_data()

    # ---------- KISAYOLLAR (TIER 2 + TIER 4) ----------

    def setup_shortcuts(self):
        QShortcut(QKeySequence("Ctrl+N"), self, self.tarif_ekle)
        QShortcut(QKeySequence("Ctrl+W"), self, self.wizard_ac)
        QShortcut(QKeySequence("Ctrl+F"), self, lambda: self.arama_input.setFocus())
        QShortcut(QKeySequence("Ctrl+D"), self, self.tarif_detay_ac)
        QShortcut(QKeySequence("Ctrl+E"), self, self.tarif_duzenle)
        QShortcut(QKeySequence("F5"), self, self.tarifleri_listele)
        QShortcut(QKeySequence("Delete"), self, self.tarif_sil)
        QShortcut(QKeySequence("Ctrl+Q"), self, self._cikis)

    def _cikis(self):
        cevap = msg_question(self, "Çıkış", "Uygulamadan çıkmak istiyor musunuz?")
        if cevap == QMessageBox.Yes:
            QApplication.quit()

    def toggle_theme(self):
        self.current_theme = "light" if self.current_theme == "dark" else "dark"
        self.setStyleSheet(ThemeManager.get(self.current_theme))
        if hasattr(self, 'theme_btn'):
            self.theme_btn.setText("☀️ AÇIK TEMA" if self.current_theme == "dark" else "🌙 KOYU TEMA")

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # HEADER
        header_widget = QFrame()
        header_widget.setStyleSheet("""
            QFrame { background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 #1a1a2e, stop:0.3 #2d2d3a, stop:0.7 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 25px; border: 1px solid #f5a623; }
        """)
        header_widget.setFixedHeight(100)
        header_layout = QHBoxLayout()
        header_layout.setContentsMargins(40, 0, 40, 0)

        logo_label = QLabel("🍽️")
        logo_label.setFont(QFont("Arial", 40))
        logo_label.setStyleSheet("color: #f5a623;")
        header_layout.addWidget(logo_label)

        title_layout = QVBoxLayout()
        title_label = QLabel("YEMEK TARİF PLATFORMU")
        title_label.setFont(QFont("Arial", 20, QFont.Bold))
        title_label.setStyleSheet("color: #f5a623;")
        title_layout.addWidget(title_label)
        subtitle_label = QLabel("LUXURY RECIPE MANAGEMENT SYSTEM — PRO VERSION")
        subtitle_label.setFont(QFont("Arial", 10))
        subtitle_label.setStyleSheet("color: #4CAF50; letter-spacing: 2px;")
        title_layout.addWidget(subtitle_label)
        header_layout.addLayout(title_layout)
        header_layout.addStretch()

        # PRO VERSION badge
        pro_badge = QLabel("⭐ PRO")
        pro_badge.setFont(QFont("Arial", 11, QFont.Bold))
        pro_badge.setStyleSheet("""
            color: #1a1a2e; background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 #f5a623, stop:1 #FFD700);
            border-radius: 12px; padding: 8px 18px; letter-spacing: 1px;
        """)
        header_layout.addWidget(pro_badge)

        # Exit butonu
        exit_btn = QPushButton("✕ EXIT")
        exit_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #e63946, stop:1 #1a1a2e);
                color: #ffffff; border: none; border-radius: 12px;
                font-weight: bold; font-size: 12px; padding: 10px 20px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #e63946, stop:1 #e63946);
            }
        """)
        exit_btn.clicked.connect(self._cikis)
        header_layout.addWidget(exit_btn)

        right_widget = QWidget()
        right_layout = QVBoxLayout()
        right_layout.setSpacing(5)
        self.istatistik_label = QLabel("🍽️ Yükleniyor...")
        self.istatistik_label.setFont(QFont("Arial", 11))
        self.istatistik_label.setStyleSheet("color: #ffffff; background-color: #2d2d3a; padding: 8px 15px; border-radius: 20px;")
        right_layout.addWidget(self.istatistik_label)
        tarih_label = QLabel(datetime.now().strftime("%d %B %Y - %H:%M"))
        tarih_label.setFont(QFont("Arial", 11))
        tarih_label.setStyleSheet("color: #f5a623; background-color: #1a1a2e; padding: 8px 15px; border-radius: 20px;")
        right_layout.addWidget(tarih_label)
        right_widget.setLayout(right_layout)
        header_layout.addWidget(right_widget)
        header_widget.setLayout(header_layout)

        # DASHBOARD KARTLARI (TIER 4: 6 kart)
        dashboard_layout = QHBoxLayout()
        dashboard_layout.setSpacing(15)
        self.tarif_card = self.create_luxury_card("📖 TOPLAM TARİFLER", "0", "#f5a623", "✨ Lezzetli Tarifler")
        self.kategori_card = self.create_luxury_card("📚 KATEGORİLER", "0", "#4CAF50", "🌸 Zengin Çeşitlilik")
        self.kullanici_card = self.create_luxury_card("👥 KULLANICILAR", "0", "#2196F3", "🌟 Aktif Kullanıcılar")
        self.puan_card = self.create_luxury_card("⭐ DEĞERLENDİRMELER", "0", "#e63946", "💬 Kullanıcı Yorumları")
        self.not_card = self.create_luxury_card("📝 KİŞİSEL NOTLAR", "0", "#9c27b0", "💡 Kişisel Gözlemler")   # TIER 4
        self.versiyon_card = self.create_luxury_card("📜 VERSİYONLAR", "0", "#00BCD4", "🔖 Tarih Arşivi")       # TIER 4
        dashboard_layout.addWidget(self.tarif_card)
        dashboard_layout.addWidget(self.kategori_card)
        dashboard_layout.addWidget(self.kullanici_card)
        dashboard_layout.addWidget(self.puan_card)
        dashboard_layout.addWidget(self.not_card)
        dashboard_layout.addWidget(self.versiyon_card)

        # MENU (TIER 4: 6 sekme)
        menu_widget = QFrame()
        menu_widget.setStyleSheet("QFrame { background-color: #2d2d3a; border-radius: 20px; }")
        menu_widget.setFixedHeight(60)
        menu_layout = QHBoxLayout()
        menu_layout.setContentsMargins(20, 0, 20, 0)
        menu_layout.setSpacing(8)

        self.menu_buttons = {}
        menu_items = [
            ("📖 TARİFLER", self.show_tarifler),
            ("👥 KULLANICILAR", self.show_kullanicilar),
            ("⭐ DEĞERLENDİRMELER", self.show_degerlendirmeler),
            ("📋 YAPILANLAR", self.show_yapilanlar),
            ("📝 NOTLARIM", self.show_notlar),         # TIER 4
            ("📊 İSTATİSTİKLER", self.show_istatistikler)
        ]
        for text, func in menu_items:
            btn = QPushButton(text)
            btn.setStyleSheet("""
                QPushButton { background-color: transparent; color: #ffffff; border: none;
                    border-radius: 15px; font-weight: bold; font-size: 12px; padding: 10px 18px; }
                QPushButton:hover { background-color: #f5a623; color: #1a1a2e; }
                QPushButton:checked { background-color: #f5a623; color: #1a1a2e; }
            """)
            btn.setCheckable(True)
            btn.clicked.connect(func)
            menu_layout.addWidget(btn)
            self.menu_buttons[text] = btn
        menu_layout.addStretch()
        menu_widget.setLayout(menu_layout)

        # STACKED WIDGET
        self.stacked_widget = QStackedWidget()
        self.stacked_widget.setStyleSheet("QStackedWidget { background-color: transparent; }")
        self.tarif_page = self.create_tarif_page()
        self.kullanici_page = self.create_kullanici_page()
        self.degerlendirme_page = self.create_degerlendirme_page()
        self.yapilan_page = self.create_yapilan_page()
        self.notlar_page = NotlarPage(self.db, self)   # TIER 4
        self.stats_page = StatisticsWidget(self.db)

        self.stacked_widget.addWidget(self.tarif_page)
        self.stacked_widget.addWidget(self.kullanici_page)
        self.stacked_widget.addWidget(self.degerlendirme_page)
        self.stacked_widget.addWidget(self.yapilan_page)
        self.stacked_widget.addWidget(self.notlar_page)
        self.stacked_widget.addWidget(self.stats_page)
        self.menu_buttons["📖 TARİFLER"].setChecked(True)

        main_layout.addWidget(header_widget)
        main_layout.addLayout(dashboard_layout)
        main_layout.addWidget(menu_widget)
        main_layout.addWidget(self.stacked_widget)
        central_widget.setLayout(main_layout)

        self.timer = QTimer()
        self.timer.timeout.connect(self.refresh_all)
        self.timer.start(5000)

    def create_luxury_card(self, title, value, color, subtitle):
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{ background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #2d2d3a, stop:1 #1a1a2e);
                border-radius: 20px; border-left: 8px solid {color}; }}
        """)
        layout = QVBoxLayout()
        layout.setSpacing(8)
        layout.setContentsMargins(18, 18, 18, 18)
        title_label = QLabel(title)
        title_label.setFont(QFont("Arial", 10, QFont.Bold))
        title_label.setStyleSheet(f"color: {color};")
        value_label = QLabel(value)
        value_label.setFont(QFont("Arial", 28, QFont.Bold))
        value_label.setStyleSheet("color: #ffffff;")
        value_label.setObjectName("value_label")
        subtitle_label = QLabel(subtitle)
        subtitle_label.setFont(QFont("Arial", 8))
        subtitle_label.setStyleSheet("color: #4CAF50;")
        layout.addWidget(title_label)
        layout.addWidget(value_label)
        layout.addWidget(subtitle_label)
        card.setLayout(layout)
        return card

    # ---------- SEKME GEÇİŞLERİ ----------

    def _menu_sec(self, aktif):
        for btn in self.menu_buttons.values():
            btn.setChecked(False)
        self.menu_buttons[aktif].setChecked(True)

    def show_tarifler(self):
        self.stacked_widget.setCurrentWidget(self.tarif_page)
        self._menu_sec("📖 TARİFLER")
        self.tarifleri_listele()

    def show_kullanicilar(self):
        self.stacked_widget.setCurrentWidget(self.kullanici_page)
        self._menu_sec("👥 KULLANICILAR")
        self.kullanicilari_listele()

    def show_degerlendirmeler(self):
        self.stacked_widget.setCurrentWidget(self.degerlendirme_page)
        self._menu_sec("⭐ DEĞERLENDİRMELER")
        self.degerlendirmeleri_listele()

    def show_yapilanlar(self):
        self.stacked_widget.setCurrentWidget(self.yapilan_page)
        self._menu_sec("📋 YAPILANLAR")
        self.yapilanlari_listele()

    def show_notlar(self):  # TIER 4
        self.stacked_widget.setCurrentWidget(self.notlar_page)
        self._menu_sec("📝 NOTLARIM")
        self.notlar_page.notlari_yukle()

    def show_istatistikler(self):
        self.stacked_widget.setCurrentWidget(self.stats_page)
        self._menu_sec("📊 İSTATİSTİKLER")
        self.stats_page.update_charts()

    # ---------- TARİF SAYFASI ----------

    def create_tarif_page(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 20)

        # 1. Satır: Temel CRUD
        btn_layout1 = QHBoxLayout()
        btn_layout1.setSpacing(10)
        for (metin, renk, func) in [
            ("➕ YENİ TARİF", "#4CAF50", self.tarif_ekle),
            ("✨ SİHİRBAZ", "#9c27b0", self.wizard_ac),     # TIER 4
            ("✏️ DÜZENLE", "#2196F3", self.tarif_duzenle),  # TIER 4
            ("🗑️ TARİF SİL", "#e63946", self.tarif_sil),
            ("🥕 MALZEME EKLE", "#2196F3", self.malzeme_ekle),
            ("⭐ DEĞERLENDİR", "#f5a623", self.degerlendir),
            ("👁️ DETAY", "#9c27b0", self.tarif_detay_ac),
            ("🔍 ARAMA", "#2196F3", self.advanced_search),
        ]:
            b = LuxuryButton(metin, renk)
            b.clicked.connect(func)
            btn_layout1.addWidget(b)
        btn_layout1.addStretch()
        layout.addLayout(btn_layout1)

        # 2. Satır: Export/Yedek araçları
        btn_layout2 = QHBoxLayout()
        btn_layout2.setSpacing(10)
        for (metin, renk, func) in [
            ("✅ YAPTIM", "#4CAF50", self.yaptim_isaretle),
            ("📜 VERSİYONLAR", "#00BCD4", self.versiyon_gecmis_ac),
            ("📋 KOPYALA", "#795548", self.tarif_kopyala),
            ("📥 CSV", "#2196F3", self.export_csv),
            ("📄 PDF", "#e63946", self.export_pdf),
            ("📊 EXCEL", "#1D6F42", self.export_xlsx),
            ("💾 YEDEKLE", "#9c27b0", self.yedekle),
            ("♻️ GERİ YÜKLE", "#795548", self.geri_yukle),
            ("🔄 YENİLE", "#4CAF50", self.tarifleri_listele),
        ]:
            b = LuxuryButton(metin, renk)
            b.clicked.connect(func)
            btn_layout2.addWidget(b)
        btn_layout2.addStretch()
        layout.addLayout(btn_layout2)

        # Hızlı arama
        arama_frame = QFrame()
        arama_frame.setStyleSheet("QFrame { background-color: #2d2d3a; border-radius: 15px; border: 1px solid #f5a623; }")
        arama_frame.setFixedHeight(54)
        arama_layout = QHBoxLayout()
        arama_layout.setContentsMargins(15, 6, 15, 6)
        arama_layout.setSpacing(10)
        arama_label = QLabel("🔍 Hızlı ara:")
        arama_label.setStyleSheet("color: #f5a623; font-weight: bold;")
        arama_layout.addWidget(arama_label)
        self.arama_input = QLineEdit()
        self.arama_input.setPlaceholderText("Tarif, kategori veya malzeme...  (Ctrl+F)")
        self.arama_input.setStyleSheet("QLineEdit { background-color: #0f0f1a; color: #ffffff; border: 1px solid #4CAF50; border-radius: 8px; padding: 8px 12px; font-size: 12px; } QLineEdit:focus { border: 2px solid #4CAF50; }")
        self.arama_input.textChanged.connect(self.real_time_ara)
        arama_layout.addWidget(self.arama_input)
        temizle_btn = QPushButton("✕")
        temizle_btn.setFixedWidth(40)
        temizle_btn.setStyleSheet("QPushButton { background-color: #e63946; color: white; border: none; border-radius: 8px; font-weight: bold; } QPushButton:hover { background-color: #d63335; }")
        temizle_btn.clicked.connect(lambda: self.arama_input.setText(""))
        arama_layout.addWidget(temizle_btn)
        arama_frame.setLayout(arama_layout)
        layout.addWidget(arama_frame)

        # Filtreler
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("📌 Kategori:"))
        self.kategori_filter = QComboBox()
        self.kategori_filter.addItem("📌 Tümü")
        self.kategori_filter.addItems(self.db.kategorileri_getir())
        self.kategori_filter.currentTextChanged.connect(self.tarifleri_listele)
        filter_layout.addWidget(self.kategori_filter)
        self.favorit_filter = QCheckBox("⭐ Sadece Favoriler")
        self.favorit_filter.stateChanged.connect(self.tarifleri_listele)
        filter_layout.addWidget(self.favorit_filter)
        ipucu = QLabel("💡 Ctrl+W=Sihirbaz · Ctrl+E=Düzenle · Ctrl+D=Detay · çift tık=detay · sağ tık=menü")
        ipucu.setStyleSheet("color: #999999; font-size: 10px;")
        filter_layout.addWidget(ipucu)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # Tablo
        self.tarif_table = QTableWidget()
        self.tarif_table.setColumnCount(7)
        self.tarif_table.setHorizontalHeaderLabels(["ID", "TARİF ADI", "KATEGORİ", "SÜRE", "ZORLUK", "PUAN", "MALİYET (₺)"])
        self.tarif_table.setAlternatingRowColors(True)
        self.tarif_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tarif_table.doubleClicked.connect(self.show_tarif_detay)
        self.tarif_table.setSortingEnabled(True)
        self.tarif_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tarif_table.customContextMenuRequested.connect(self.tarif_context_menu)
        layout.addWidget(self.tarif_table)
        widget.setLayout(layout)
        return widget

    # ---------- SAĞ TIK MENÜSÜ (TIER 2 + TIER 4) ----------

    def tarif_context_menu(self, pozisyon):
        row = self.tarif_table.currentRow()
        if row < 0:
            return
        menu = QMenu(self)
        for metin, func in [
            ("👁️ Detayları Gör", self.tarif_detay_ac),
            ("✏️ Düzenle", self.tarif_duzenle),           # TIER 4
            ("📜 Versiyon Geçmişi", self.versiyon_gecmis_ac),  # TIER 4
            ("⭐ Favoriye Ekle / Çıkar", self.favori_toggle_secili),
            ("✅ Yaptım Olarak İşaretle", self.yaptim_isaretle),
            ("🥕 Malzeme Ekle", self.malzeme_ekle),
        ]:
            act = QAction(metin, self)
            act.triggered.connect(func)
            menu.addAction(act)
        menu.addSeparator()
        sil_act = QAction("🗑️ Sil", self)
        sil_act.triggered.connect(self.tarif_sil)
        menu.addAction(sil_act)
        menu.exec_(self.tarif_table.viewport().mapToGlobal(pozisyon))

    def favori_toggle_secili(self):
        row = self.tarif_table.currentRow()
        if row >= 0:
            tarif_id = int(self.tarif_table.item(row, 0).text())
            t = self.db.tarif_bul(tarif_id)
            self.db.tarif_favoriye_ekle(tarif_id, not t['is_favorite'])
            self.tarifleri_listele()
            self.update_dashboard()

    # ---------- DİĞER SAYFALAR ----------

    def create_kullanici_page(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(15)
        for metin, renk, func in [
            ("➕ YENİ KULLANICI", "#4CAF50", self.kullanici_ekle),
            ("🗑️ KULLANICI SİL", "#e63946", self.kullanici_sil),
            ("🔄 YENİLE", "#4CAF50", self.kullanicilari_listele),
        ]:
            b = LuxuryButton(metin, renk)
            b.clicked.connect(func)
            btn_layout.addWidget(b)
        btn_layout.addStretch()
        self.kullanici_table = QTableWidget()
        self.kullanici_table.setColumnCount(4)
        self.kullanici_table.setHorizontalHeaderLabels(["ID", "AD", "SOYAD", "EMAIL"])
        self.kullanici_table.setAlternatingRowColors(True)
        self.kullanici_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.kullanici_table.setSortingEnabled(True)
        layout.addLayout(btn_layout)
        layout.addWidget(self.kullanici_table)
        widget.setLayout(layout)
        return widget

    def create_degerlendirme_page(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        yenile_btn = LuxuryButton("🔄 YENİLE", "#4CAF50")
        yenile_btn.clicked.connect(self.degerlendirmeleri_listele)
        self.degerlendirme_table = QTableWidget()
        self.degerlendirme_table.setColumnCount(5)
        self.degerlendirme_table.setHorizontalHeaderLabels(["TARİF", "KULLANICI", "PUAN", "YORUM", "TARİH"])
        self.degerlendirme_table.setAlternatingRowColors(True)
        self.degerlendirme_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.degerlendirme_table.setSortingEnabled(True)
        layout.addWidget(yenile_btn)
        layout.addWidget(self.degerlendirme_table)
        widget.setLayout(layout)
        return widget

    def create_yapilan_page(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        bilgi = QLabel("📋 Yaptığın tarifler burada listelenir. 'Yaptım' işareti koymak için Tarifler sekmesini kullan.")
        bilgi.setStyleSheet("color: #f5a623; font-weight: bold;")
        layout.addWidget(bilgi)
        yenile_btn = LuxuryButton("🔄 YENİLE", "#4CAF50")
        yenile_btn.clicked.connect(self.yapilanlari_listele)
        layout.addWidget(yenile_btn)
        self.yapilan_table = QTableWidget()
        self.yapilan_table.setColumnCount(3)
        self.yapilan_table.setHorizontalHeaderLabels(["TARİF", "YAPILMA TARİHİ", "NOTLAR"])
        self.yapilan_table.setAlternatingRowColors(True)
        self.yapilan_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.yapilan_table.setSortingEnabled(True)
        layout.addWidget(self.yapilan_table)
        widget.setLayout(layout)
        return widget

    def yapilanlari_listele(self):
        self.yapilan_table.setSortingEnabled(False)
        self.yapilan_table.setRowCount(0)
        for log in self.db.yapilmis_loglari_getir():
            row = self.yapilan_table.rowCount()
            self.yapilan_table.insertRow(row)
            self.yapilan_table.setItem(row, 0, QTableWidgetItem(log['ad']))
            self.yapilan_table.setItem(row, 1, QTableWidgetItem(str(log['yapilma_tarihi'])[:16]))
            self.yapilan_table.setItem(row, 2, QTableWidgetItem(log['notlar'] or '-'))
        self.yapilan_table.setSortingEnabled(True)

    # ---------- VERİ YÜKLEME ----------

    def closeEvent(self, event):
        cevap = msg_question(self, "Çıkış", "Uygulamadan çıkmak istiyor musunuz?")
        if cevap == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()

    def load_data(self):
        self.tarifleri_listele()
        self.kullanicilari_listele()
        self.degerlendirmeleri_listele()
        self.yapilanlari_listele()
        self.update_dashboard()

    def refresh_all(self):
        self.tarifleri_listele()
        self.kullanicilari_listele()
        self.degerlendirmeleri_listele()
        self.update_dashboard()
        w = self.stacked_widget.currentWidget()
        if w == self.stats_page:
            self.stats_page.update_charts()
        elif w == self.yapilan_page:
            self.yapilanlari_listele()
        elif w == self.notlar_page:
            self.notlar_page.notlari_yukle()

    def update_dashboard(self):
        ist = self.db.istatistikler()
        self.tarif_card.findChild(QLabel, "value_label").setText(str(ist["toplam_tarif"]))
        self.kategori_card.findChild(QLabel, "value_label").setText(str(ist["toplam_kategori"]))
        self.kullanici_card.findChild(QLabel, "value_label").setText(str(ist["toplam_kullanici"]))
        self.puan_card.findChild(QLabel, "value_label").setText(str(ist["toplam_degerlendirme"]))
        self.not_card.findChild(QLabel, "value_label").setText(str(ist["toplam_not"]))
        self.versiyon_card.findChild(QLabel, "value_label").setText(str(ist["toplam_versiyon"]))
        self.istatistik_label.setText(
            f"🍽️ {ist['toplam_tarif']} Tarif | ⭐ {ist['favoriler_sayi']} Favori | "
            f"✅ {ist['toplam_yapilmis']} Yapılan | 📝 {ist['toplam_not']} Not | 👥 {ist['toplam_kullanici']} Kullanıcı"
        )

    # ---------- ARAMA ----------

    def get_zorluk_etiketi(self, z):
        return {1: "👶 Çok Kolay", 2: "😊 Kolay", 3: "😐 Orta", 4: "🤓 Zor", 5: "🔥 Çok Zor"}.get(z, "Belirsiz")

    def advanced_search(self):
        dialog = AramaDialog(self.db, self)
        if dialog.exec_() == QDialog.Accepted and dialog.results is not None:
            self.display_search_results(dialog.results)

    def real_time_ara(self):
        arama_terimi = self.arama_input.text().strip()
        if arama_terimi:
            self.display_search_results(self.db.tarif_ara(arama_terimi))
        else:
            self.tarifleri_listele()

    def display_search_results(self, tarifler):
        self.tarif_table.setSortingEnabled(False)
        self.tarif_table.setRowCount(0)
        for t in tarifler:
            row = self.tarif_table.rowCount()
            self.tarif_table.insertRow(row)
            puan = self.db.ortalama_puan(t['tarif_id'])
            maliyet = self.db.tahmini_tarif_maliyeti(t['tarif_id'])
            self.tarif_table.setItem(row, 0, NumericTableWidgetItem(str(t['tarif_id']), t['tarif_id']))
            ad_item = QTableWidgetItem(("⭐ " if t.get('is_favorite') else "") + t['ad'])
            if t.get('is_favorite'):
                ad_item.setForeground(QColor("#f5a623"))
            self.tarif_table.setItem(row, 1, ad_item)
            self.tarif_table.setItem(row, 2, QTableWidgetItem(t['kategori']))
            self.tarif_table.setItem(row, 3, NumericTableWidgetItem(f"{t['hazirlama_suresi']} dk", t['hazirlama_suresi']))
            zorluk_item = NumericTableWidgetItem(self.get_zorluk_etiketi(t.get('zorluk_seviyesi', 3)), t.get('zorluk_seviyesi', 3))
            zorluk_item.setForeground(QColor("#2196F3"))
            self.tarif_table.setItem(row, 4, zorluk_item)
            puan_item = NumericTableWidgetItem(f"⭐ {puan}" if puan > 0 else "⭐ Yok", puan)
            puan_item.setForeground(QColor("#4CAF50" if puan >= 4 else "#f5a623" if puan >= 3 else "#e63946"))
            self.tarif_table.setItem(row, 5, puan_item)
            maliyet_item = NumericTableWidgetItem(f"₺{maliyet:.2f}", maliyet)
            maliyet_item.setForeground(QColor("#4CAF50"))
            self.tarif_table.setItem(row, 6, maliyet_item)
        self.tarif_table.setSortingEnabled(True)

    def tarifleri_listele(self):
        kategori = self.kategori_filter.currentText()
        if self.arama_input.text().strip():
            tarifler = self.db.tarif_ara(self.arama_input.text().strip())
        else:
            tarifler = self.db.tarifleri_getir(kategori if kategori != "📌 Tümü" else None)
        if self.favorit_filter.isChecked():
            tarifler = [t for t in tarifler if t.get('is_favorite')]
        self.display_search_results(tarifler)

    # ---------- DETAY ----------

    def show_tarif_detay(self, index):
        if index.isValid():
            try:
                tarif_id = int(self.tarif_table.item(index.row(), 0).text())
                dialog = TarifDetayDialog(self.db, tarif_id, self)
                dialog.exec_()
                self.tarifleri_listele()
                self.update_dashboard()
                self.yapilanlari_listele()
            except (ValueError, AttributeError):
                pass

    def tarif_detay_ac(self):
        row = self.tarif_table.currentRow()
        if row >= 0:
            tarif_id = int(self.tarif_table.item(row, 0).text())
            dialog = TarifDetayDialog(self.db, tarif_id, self)
            dialog.exec_()
            self.tarifleri_listele()
            self.update_dashboard()
            self.yapilanlari_listele()
        else:
            msg_warn(self, "Uyarı", "Lütfen bir tarif seçin!")

    # ---------- TIER 4: SİHİRBAZ ----------

    def wizard_ac(self):
        wizard = TarifWizard(self.db, self)
        if wizard.exec_() == QDialog.Accepted:
            self.tarifleri_listele()
            self.update_dashboard()
            self.stats_page.update_charts()

    # ---------- TIER 4: DÜZENLE ----------

    def tarif_duzenle(self):
        row = self.tarif_table.currentRow()
        if row < 0:
            msg_warn(self, "Uyarı", "Lütfen düzenlemek istediğiniz tarifi seçin!")
            return
        tarif_id = int(self.tarif_table.item(row, 0).text())
        dialog = TarifDuzenleDialog(self.db, tarif_id, self)
        if dialog.exec_() == QDialog.Accepted and dialog.result:
            r = dialog.result
            self.db.tarif_guncelle(tarif_id, r['ad'], r['kategori'], r['hazirlama_suresi'],
                                   r['talimat'], r['zorluk_seviyesi'], r['degisiklik_notu'])
            msg_info(self, "Başarılı", "✅ Tarif güncellendi ve versiyonu kaydedildi!")
            self.tarifleri_listele()
            self.update_dashboard()
            self.stats_page.update_charts()

    # ---------- TIER 4: VERSİYON GEÇMİŞİ ----------

    def versiyon_gecmis_ac(self):
        row = self.tarif_table.currentRow()
        if row < 0:
            msg_warn(self, "Uyarı", "Lütfen bir tarif seçin!")
            return
        tarif_id = int(self.tarif_table.item(row, 0).text())
        dialog = VersiyonGecmisDialog(self.db, tarif_id, self)
        if dialog.exec_() == QDialog.Accepted:
            self.tarifleri_listele()

    # ---------- YAPTIM ----------

    def yaptim_isaretle(self):
        row = self.tarif_table.currentRow()
        if row >= 0:
            tarif_id = int(self.tarif_table.item(row, 0).text())
            tarif_adi = self.tarif_table.item(row, 1).text().replace("⭐ ", "")
            self.db.yapilmis_log_ekle(tarif_id, "")
            msg_info(self, "Başarılı", f"✅ '{tarif_adi}' yapıldı olarak kaydedildi!")
            self.update_dashboard()
            self.yapilanlari_listele()
        else:
            msg_warn(self, "Uyarı", "Lütfen bir tarif seçin!")

    # ---------- CSV EXPORT ----------

    def export_csv(self):
        yol, _ = QFileDialog.getSaveFileName(self, "CSV Olarak Kaydet", "tarifler.csv", "CSV Dosyaları (*.csv)")
        if not yol:
            return
        try:
            tarifler = self.db.tarifleri_getir()
            with open(yol, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(["ID", "Ad", "Kategori", "Süre (dk)", "Zorluk", "Puan",
                                 "Maliyet (TL)", "Kalori", "Protein (g)", "Favori"])
                for t in tarifler:
                    writer.writerow([
                        t['tarif_id'], t['ad'], t['kategori'], t['hazirlama_suresi'],
                        self.get_zorluk_etiketi(t['zorluk_seviyesi']),
                        self.db.ortalama_puan(t['tarif_id']),
                        self.db.tahmini_tarif_maliyeti(t['tarif_id']),
                        self.db.tahmini_tarif_kalorileri(t['tarif_id']),
                        self.db.tahmini_tarif_proteini(t['tarif_id']),
                        "Evet" if t['is_favorite'] else "Hayır"
                    ])
            msg_info(self, "Başarılı", f"✅ {len(tarifler)} tarif CSV olarak kaydedildi:\n{yol}")
        except Exception as e:
            msg_warn(self, "Hata", f"CSV kaydedilemedi:\n{e}")

    # ---------- XLSX EXPORT ----------

    def export_xlsx(self):
        """Tüm tarifleri Excel formatında dışa aktar."""
        if not OPENPYXL_VAR:
            msg_warn(self, "Hata",
                "openpyxl kurulu değil!\n\nKurmak için:\npip install openpyxl")
            return

        yol, _ = QFileDialog.getSaveFileName(
            self, "Excel Olarak Kaydet", "tarifler.xlsx",
            "Excel Dosyaları (*.xlsx)"
        )
        if not yol:
            return
        try:
            wb = openpyxl.Workbook()

            # ── Sayfa 1: Tarifler ──
            ws1 = wb.active
            ws1.title = "Tarifler"
            baslik_font = Font(bold=True, color="FFFFFF", size=11)
            altin_fill = PatternFill("solid", fgColor="f5a623")
            orta = Alignment(horizontal="center", vertical="center")
            ince = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            basliklar = ["ID", "Tarif Adı", "Kategori", "Süre (dk)",
                         "Zorluk", "Ort. Puan", "Maliyet (₺)",
                         "Kalori", "Protein (g)", "Favori"]
            ws1.append(basliklar)
            for col, _ in enumerate(basliklar, 1):
                c = ws1.cell(1, col)
                c.font = baslik_font
                c.fill = altin_fill
                c.alignment = orta
                c.border = ince

            tarifler = self.db.tarifleri_getir()
            for t in tarifler:
                ws1.append([
                    t['tarif_id'], t['ad'], t['kategori'],
                    t['hazirlama_suresi'],
                    self.get_zorluk_etiketi(t['zorluk_seviyesi']),
                    self.db.ortalama_puan(t['tarif_id']),
                    self.db.tahmini_tarif_maliyeti(t['tarif_id']),
                    self.db.tahmini_tarif_kalorileri(t['tarif_id']),
                    self.db.tahmini_tarif_proteini(t['tarif_id']),
                    "Evet" if t['is_favorite'] else "Hayır"
                ])

            for row_idx, row in enumerate(ws1.iter_rows(min_row=2), 2):
                fill = PatternFill("solid", fgColor="2d2d3a") if row_idx % 2 == 0 \
                       else PatternFill("solid", fgColor="1a1a2e")
                for cell in row:
                    cell.fill = fill
                    cell.font = Font(color="FFFFFF")
                    cell.border = ince
                    cell.alignment = orta

            for i, col in enumerate(ws1.columns, 1):
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws1.column_dimensions[get_column_letter(i)].width = min(max(max_len + 3, 10), 35)

            # ── Sayfa 2: Malzemeler ──
            ws2 = wb.create_sheet("Malzemeler")
            ws2.append(["Tarif ID", "Tarif Adı", "Malzeme", "Miktar", "Birim", "Fiyat (₺)", "Kalori", "Protein"])
            for col in range(1, 9):
                c = ws2.cell(1, col)
                c.font = baslik_font; c.fill = altin_fill
                c.alignment = orta; c.border = ince

            for t in tarifler:
                for m in self.db.malzemeleri_getir(t['tarif_id']):
                    ws2.append([
                        t['tarif_id'], t['ad'], m['malzeme_adi'],
                        m['miktar'], m.get('birim', 'gr'),
                        m.get('fiyat', 0), m.get('kalori', 0), m.get('protein', 0)
                    ])

            for i, col in enumerate(ws2.columns, 1):
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws2.column_dimensions[get_column_letter(i)].width = min(max(max_len + 3, 10), 35)

            # ── Sayfa 3: Değerlendirmeler ──
            ws3 = wb.create_sheet("Degerlendirmeler")
            ws3.append(["Tarif", "Kullanıcı", "Puan", "Yorum", "Tarih"])
            for col in range(1, 6):
                c = ws3.cell(1, col)
                c.font = baslik_font; c.fill = altin_fill
                c.alignment = orta; c.border = ince

            for t in tarifler:
                for d in self.db.degerlendirmeleri_getir(t['tarif_id']):
                    ws3.append([
                        t['ad'], f"{d['ad']} {d['soyad']}",
                        d['puan'], d['yorum'] or '', str(d['tarih'])[:16]
                    ])

            wb.save(yol)
            msg_info(self, "Başarılı",
                f"✅ {len(tarifler)} tarif Excel'e kaydedildi:\n{yol}\n\n"
                "📋 Sayfalar: Tarifler | Malzemeler | Değerlendirmeler")

        except Exception as e:
            msg_warn(self, "Hata", f"Excel kaydedilemedi:\n{e}")

    # ---------- TARİF KOPYALA ----------

    def tarif_kopyala(self):
        row = self.tarif_table.currentRow()
        if row < 0:
            msg_warn(self, "Uyarı", "Lütfen bir tarif seçin!")
            return
        tarif_id = int(self.tarif_table.item(row, 0).text())
        tarif_adi = self.tarif_table.item(row, 1).text().replace("⭐ ", "")
        reply = msg_question(self, "Tarif Kopyala", f"'{tarif_adi}' tarifini kopyalamak istiyor musunuz?\n"
                                     "Tüm malzemeleriyle birlikte yeni tarif oluşturulacak.")
        if reply == QMessageBox.Yes:
            yeni_id = self.db.tarif_kopyala(tarif_id)
            if yeni_id:
                msg_info(self, "Başarılı",
                    f"✅ '{tarif_adi}' kopyalandı!\nYeni tarif: '{tarif_adi} (Kopya)'")
                self.tarifleri_listele()
                self.update_dashboard()
                self.stats_page.update_charts()

    def export_pdf(self):
        row = self.tarif_table.currentRow()
        if row < 0:
            msg_warn(self, "Uyarı", "Lütfen bir tarif seçin!")
            return
        tarif_id = int(self.tarif_table.item(row, 0).text())
        tarif = self.db.tarif_bul(tarif_id)
        malzemeler = self.db.malzemeleri_getir(tarif_id)
        if not malzemeler:
            msg_warn(self, "Uyarı", "Bu tarifin malzemesi yok!")
            return
        yol, _ = QFileDialog.getSaveFileName(self, "PDF Olarak Kaydet",
                                             f"alisveris_{tarif['ad']}.pdf", "PDF Dosyaları (*.pdf)")
        if not yol:
            return
        try:
            from matplotlib.backends.backend_pdf import PdfPages
            import matplotlib.pyplot as plt
            satirlar = [["Malzeme", "Miktar", "Birim", "Fiyat (TL)"]]
            toplam = 0.0
            for m in malzemeler:
                fiyat = m.get('fiyat', 0) or 0
                toplam += fiyat
                satirlar.append([m['malzeme_adi'], str(m['miktar']), m.get('birim', 'gr'), f"{fiyat:.2f}"])
            satirlar.append(["", "", "TOPLAM", f"{toplam:.2f} TL"])
            with PdfPages(yol) as pdf:
                fig, ax = plt.subplots(figsize=(8, max(4, len(satirlar) * 0.4 + 1)))
                ax.axis('off')
                ax.set_title(f"Alisveris Listesi: {tarif['ad']}\n({tarif['kategori']} - {tarif['hazirlama_suresi']} dk)",
                             fontsize=14, fontweight='bold', pad=20)
                tablo = ax.table(cellText=satirlar, loc='center', cellLoc='left')
                tablo.auto_set_font_size(False)
                tablo.set_fontsize(10)
                tablo.scale(1, 1.6)
                for j in range(4):
                    tablo[0, j].set_facecolor('#f5a623')
                    tablo[0, j].set_text_props(weight='bold', color='white')
                    tablo[len(satirlar)-1, j].set_facecolor('#4CAF50')
                    tablo[len(satirlar)-1, j].set_text_props(weight='bold', color='white')
                pdf.savefig(fig, bbox_inches='tight')
                plt.close(fig)
            msg_info(self, "Başarılı", f"✅ Alışveriş listesi kaydedildi:\n{yol}")
        except Exception as e:
            msg_warn(self, "Hata", f"PDF oluşturulamadı:\n{e}")

    # ---------- YEDEKLEME ----------

    def yedekle(self):
        yol, _ = QFileDialog.getSaveFileName(self, "Veritabanını Yedekle",
                                             f"yemek_tarif_yedek_{datetime.now().strftime('%Y%m%d_%H%M')}.db",
                                             "Veritabanı (*.db)")
        if not yol:
            return
        try:
            self.db.yedekle(yol)
            msg_info(self, "Başarılı", f"✅ Veritabanı yedeklendi:\n{yol}")
        except Exception as e:
            msg_warn(self, "Hata", f"Yedekleme başarısız:\n{e}")

    def geri_yukle(self):
        yol, _ = QFileDialog.getOpenFileName(self, "Yedekten Geri Yükle", "", "Veritabanı (*.db)")
        if not yol:
            return
        reply = msg_question(self, "Onay", "⚠️ Mevcut veritabanı yedekteki veriyle DEĞİŞTİRİLECEK!\nDevam etmek istiyor musunuz?")
        if reply == QMessageBox.Yes:
            try:
                self.db.geri_yukle(yol)
                msg_info(self, "Başarılı", "✅ Geri yükleme tamamlandı!")
                self.load_data()
            except Exception as e:
                msg_warn(self, "Hata", f"Geri yükleme başarısız:\n{e}")

    # ---------- CRUD: TARİF ----------

    def tarif_ekle(self):
        dialog = TarifEkleDialog(self.db, self)
        if dialog.exec_() == QDialog.Accepted and dialog.result:
            self.db.tarif_ekle(*dialog.result)
            msg_info(self, "Başarılı", "✅ Tarif başarıyla eklendi!")
            self.tarifleri_listele()
            self.update_dashboard()
            self.stats_page.update_charts()

    def tarif_sil(self):
        row = self.tarif_table.currentRow()
        if row >= 0:
            tarif_id = int(self.tarif_table.item(row, 0).text())
            tarif_adi = self.tarif_table.item(row, 1).text()
            reply = msg_question(self, "Silme Onayı", f"'{tarif_adi}' tarifini silmek istediğinize emin misiniz?")
            if reply == QMessageBox.Yes:
                self.db.tarif_sil(tarif_id)
                msg_info(self, "Başarılı", "✅ Tarif silindi!")
                self.tarifleri_listele()
                self.update_dashboard()
                self.stats_page.update_charts()
        else:
            msg_warn(self, "Uyarı", "Lütfen silinecek tarifi seçin!")

    def malzeme_ekle(self):
        row = self.tarif_table.currentRow()
        if row >= 0:
            tarif_id = int(self.tarif_table.item(row, 0).text())
            tarif_adi = self.tarif_table.item(row, 1).text()
            dialog = MalzemeEkleDialog(self.db, tarif_id, tarif_adi, self)
            if dialog.exec_() == QDialog.Accepted and dialog.result:
                self.db.malzeme_ekle(*dialog.result)
                msg_info(self, "Başarılı", "✅ Malzeme başarıyla eklendi!")
                self.tarifleri_listele()
        else:
            msg_warn(self, "Uyarı", "Lütfen bir tarif seçin!")

    # ---------- CRUD: KULLANICI ----------

    def kullanicilari_listele(self):
        self.kullanici_table.setSortingEnabled(False)
        kullanicilar = self.db.kullanicilari_getir()
        self.kullanici_table.setRowCount(0)
        for k in kullanicilar:
            row = self.kullanici_table.rowCount()
            self.kullanici_table.insertRow(row)
            self.kullanici_table.setItem(row, 0, NumericTableWidgetItem(str(k['kullanici_id']), k['kullanici_id']))
            self.kullanici_table.setItem(row, 1, QTableWidgetItem(k['ad']))
            self.kullanici_table.setItem(row, 2, QTableWidgetItem(k['soyad']))
            self.kullanici_table.setItem(row, 3, QTableWidgetItem(k['email'] or '-'))
        self.kullanici_table.setSortingEnabled(True)

    def kullanici_ekle(self):
        dialog = KullaniciEkleDialog(self)
        if dialog.exec_() == QDialog.Accepted and dialog.result:
            try:
                self.db.kullanici_ekle(*dialog.result)
                msg_info(self, "Başarılı", "✅ Kullanıcı başarıyla eklendi!")
                self.kullanicilari_listele()
                self.update_dashboard()
                self.stats_page.update_charts()
            except sqlite3.IntegrityError:
                msg_warn(self, "Hata", "Bu email adresi zaten kayıtlı!")

    def kullanici_sil(self):
        row = self.kullanici_table.currentRow()
        if row >= 0:
            kullanici_id = int(self.kullanici_table.item(row, 0).text())
            kullanici_adi = f"{self.kullanici_table.item(row, 1).text()} {self.kullanici_table.item(row, 2).text()}"
            reply = msg_question(self, "Silme Onayı", f"'{kullanici_adi}' kullanıcısını silmek istediğinize emin misiniz?")
            if reply == QMessageBox.Yes:
                self.db.kullanici_sil(kullanici_id)
                msg_info(self, "Başarılı", "✅ Kullanıcı silindi!")
                self.kullanicilari_listele()
                self.degerlendirmeleri_listele()
                self.update_dashboard()
                self.stats_page.update_charts()
        else:
            msg_warn(self, "Uyarı", "Lütfen silinecek kullanıcıyı seçin!")

    # ---------- DEĞERLENDİRME ----------

    def degerlendir(self):
        row = self.tarif_table.currentRow()
        if row >= 0:
            tarif_id = int(self.tarif_table.item(row, 0).text())
            tarif_adi = self.tarif_table.item(row, 1).text().replace("⭐ ", "")
            dialog = DegerlendirmeDialog(self.db, tarif_id, tarif_adi, self)
            if dialog.exec_() == QDialog.Accepted and dialog.result:
                self.db.degerlendirme_ekle(*dialog.result)
                msg_info(self, "Başarılı", "✅ Değerlendirme kaydedildi!")
                self.tarifleri_listele()
                self.degerlendirmeleri_listele()
                self.kullanicilari_listele()
                self.update_dashboard()
                self.stats_page.update_charts()
        else:
            msg_warn(self, "Uyarı", "Lütfen bir tarif seçin!")

    def degerlendirmeleri_listele(self):
        self.degerlendirme_table.setSortingEnabled(False)
        self.degerlendirme_table.setRowCount(0)
        tarifler = self.db.tarifleri_getir()
        for tarif in tarifler:
            for d in self.db.degerlendirmeleri_getir(tarif['tarif_id']):
                row = self.degerlendirme_table.rowCount()
                self.degerlendirme_table.insertRow(row)
                self.degerlendirme_table.setItem(row, 0, QTableWidgetItem(tarif['ad']))
                self.degerlendirme_table.setItem(row, 1, QTableWidgetItem(f"{d['ad']} {d['soyad']}"))
                puan_item = NumericTableWidgetItem(f"{'⭐' * d['puan']} ({d['puan']}/5)", d['puan'])
                puan_item.setForeground(QColor("#4CAF50" if d['puan'] >= 4 else "#f5a623" if d['puan'] >= 3 else "#e63946"))
                self.degerlendirme_table.setItem(row, 2, puan_item)
                self.degerlendirme_table.setItem(row, 3, QTableWidgetItem(
                    d['yorum'][:40] + "..." if d['yorum'] and len(d['yorum']) > 40 else d['yorum'] or '-'))
                self.degerlendirme_table.setItem(row, 4, QTableWidgetItem(d['tarih'][:16]))
        self.degerlendirme_table.setSortingEnabled(True)


# ===================== MAIN =====================

def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = YemekTarifPlatformu()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()